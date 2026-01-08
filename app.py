"""SignMaker Web App - Hosted signage product generator."""
import os
import json
from pathlib import Path
from datetime import datetime

from flask import Flask, render_template_string, jsonify, request, Response, send_file
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

from config import SECRET_KEY, SIZES, COLORS, BRAND_NAME
from models import init_db, Product
from jobs import submit_job, get_job, get_all_jobs, job_to_dict, start_workers

app = Flask(__name__)
app.secret_key = SECRET_KEY

# Initialize database on startup
init_db()

# Start background workers
start_workers()

# HTML Template
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SignMaker</title>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { 
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #f5f5f5;
            min-height: 100vh;
        }
        .header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            text-align: center;
        }
        .header h1 { font-size: 2rem; margin-bottom: 5px; }
        .header p { opacity: 0.9; }
        .container { max-width: 1400px; margin: 0 auto; padding: 20px; }
        .tabs {
            display: flex;
            gap: 10px;
            margin-bottom: 20px;
            flex-wrap: wrap;
        }
        .tab {
            padding: 12px 24px;
            background: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-size: 1rem;
            transition: all 0.2s;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .tab:hover { transform: translateY(-2px); box-shadow: 0 4px 8px rgba(0,0,0,0.15); }
        .tab.active { background: #667eea; color: white; }
        .panel { display: none; }
        .panel.active { display: block; }
        .card {
            background: white;
            border-radius: 12px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }
        .card h2 { margin-bottom: 15px; color: #333; }
        table { width: 100%; border-collapse: collapse; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #eee; }
        th { background: #f8f9fa; font-weight: 600; }
        tr:hover { background: #f8f9fa; }
        .btn {
            padding: 10px 20px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 0.9rem;
            transition: all 0.2s;
        }
        .btn-primary { background: #667eea; color: white; }
        .btn-primary:hover { background: #5a6fd6; }
        .btn-success { background: #28a745; color: white; }
        .btn-success:hover { background: #218838; }
        .btn-danger { background: #dc3545; color: white; }
        .btn-danger:hover { background: #c82333; }
        .btn-secondary { background: #6c757d; color: white; }
        .btn-secondary:hover { background: #5a6268; }
        input, select, textarea {
            padding: 10px;
            border: 1px solid #ddd;
            border-radius: 6px;
            font-size: 1rem;
            width: 100%;
        }
        input:focus, select:focus, textarea:focus {
            outline: none;
            border-color: #667eea;
        }
        .form-group { margin-bottom: 15px; }
        .form-group label { display: block; margin-bottom: 5px; font-weight: 500; }
        .form-row { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; }
        .status-badge {
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 0.8rem;
            font-weight: 500;
        }
        .status-pending { background: #ffc107; color: #000; }
        .status-approved { background: #28a745; color: white; }
        .status-rejected { background: #dc3545; color: white; }
        .output-box {
            background: #1e1e1e;
            color: #0f0;
            padding: 15px;
            border-radius: 8px;
            font-family: monospace;
            font-size: 0.85rem;
            max-height: 400px;
            overflow-y: auto;
            white-space: pre-wrap;
        }
        .product-grid {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 20px;
        }
        @media (max-width: 1200px) {
            .product-grid { grid-template-columns: repeat(2, 1fr); }
        }
        @media (max-width: 800px) {
            .product-grid { grid-template-columns: 1fr; }
        }
        .product-card {
            background: white;
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }
        .product-card img {
            width: 100%;
            height: 200px;
            object-fit: contain;
            background: #f8f9fa;
        }
        .product-card-body { padding: 15px; }
        .product-card h3 { font-size: 1.1rem; margin-bottom: 10px; }
        .product-meta { color: #666; font-size: 0.9rem; }
        .actions { display: flex; gap: 10px; margin-top: 15px; }
        .actions .btn { flex: 1; }
        .alert {
            padding: 15px;
            border-radius: 8px;
            margin-bottom: 20px;
        }
        .alert-info { background: #e7f3ff; color: #0066cc; }
        .alert-success { background: #d4edda; color: #155724; }
        .alert-error { background: #f8d7da; color: #721c24; }
    </style>
</head>
<body>
    <div class="header">
        <h1>🪧 SignMaker</h1>
        <p>Signage Product Generator & Publisher</p>
    </div>
    
    <div class="container">
        <div class="tabs">
            <button class="tab active" onclick="showTab('products')">📦 Products</button>
            <button class="tab" onclick="showTab('qa')">✅ QA Review</button>
            <button class="tab" onclick="showTab('generate')">🎨 Generate</button>
            <button class="tab" onclick="showTab('export')">📤 Export</button>
        </div>
        
        <!-- Products Tab -->
        <div id="products-panel" class="panel active">
            <div class="card">
                <h2>Product List</h2>
                <div style="margin-bottom: 15px;">
                    <button class="btn btn-primary" onclick="showAddProduct()">+ Add Product</button>
                    <button class="btn btn-secondary" onclick="loadProducts()">↻ Refresh</button>
                </div>
                <table id="products-table">
                    <thead>
                        <tr>
                            <th>M Number</th>
                            <th>Description</th>
                            <th>Size</th>
                            <th>Color</th>
                            <th>EAN</th>
                            <th>Status</th>
                            <th>Actions</th>
                        </tr>
                    </thead>
                    <tbody id="products-tbody"></tbody>
                </table>
            </div>
        </div>
        
        <!-- QA Review Tab -->
        <div id="qa-panel" class="panel">
            <div class="card">
                <h2>QA Review</h2>
                <div style="display: flex; gap: 10px; align-items: center; margin-bottom: 15px;">
                    <button class="btn btn-secondary" onclick="loadQAProducts()">↻ Refresh</button>
                </div>
                <div id="qa-grid" class="product-grid" style="margin-top: 20px;"></div>
            </div>
            <!-- Debug Terminal -->
            <div class="card" style="margin-top: 20px;">
                <h3 style="margin-bottom: 10px;">Debug Console</h3>
                <div id="debug-console" class="output-box" style="height: 150px; font-size: 11px;"></div>
            </div>
        </div>
        
        <!-- Generate Tab -->
        <div id="generate-panel" class="panel">
            <div class="card">
                <h2>Generate Images & Content</h2>
                
                <!-- Step 1: Product Summary -->
                <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                    <h3 style="margin: 0 0 10px 0;">📋 Product Summary</h3>
                    <div id="product-summary" style="font-size: 13px; color: #555;">
                        <em>Loading product summary...</em>
                    </div>
                </div>
                
                <!-- Step 2: Sample Images for AI Context -->
                <div style="background: #f0f7ff; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                    <h3 style="margin: 0 0 10px 0;">🖼️ Sample Images for AI Context</h3>
                    <p style="font-size: 12px; color: #666; margin-bottom: 10px;">
                        These images will be sent to the AI to help it understand your product range.
                    </p>
                    <div id="sample-images" style="display: flex; gap: 10px; flex-wrap: wrap;">
                        <em style="color: #888;">Loading sample images...</em>
                    </div>
                </div>
                
                <!-- Step 3: AI Analysis - Auto-populate fields -->
                <div style="background: #fff3cd; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                    <h3 style="margin: 0 0 10px 0;">🤖 Step 1: Analyze Products</h3>
                    <p style="font-size: 12px; color: #666; margin-bottom: 10px;">
                        Click to have AI analyze the sample images and auto-populate the fields below.
                    </p>
                    <button class="btn btn-warning" onclick="analyzeProducts()" id="btn-analyze">🔍 Analyze Products with AI</button>
                    <span id="analyze-status" style="margin-left: 10px; font-size: 12px;"></span>
                </div>
                
                <!-- Step 4: Editable fields (populated by AI, editable by human) -->
                <div style="background: #d4edda; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                    <h3 style="margin: 0 0 10px 0;">✏️ Step 2: Review & Edit</h3>
                    <p style="font-size: 12px; color: #666; margin-bottom: 10px;">
                        Review and edit the AI-generated descriptions before generating content.
                    </p>
                    <div class="form-row">
                        <div class="form-group">
                            <label>Product Theme (for AI)</label>
                            <textarea id="theme" rows="3" placeholder="Click 'Analyze Products' to auto-populate..."></textarea>
                        </div>
                        <div class="form-group">
                            <label>Target Use Cases</label>
                            <textarea id="use-cases" rows="3" placeholder="Click 'Analyze Products' to auto-populate..."></textarea>
                        </div>
                    </div>
                </div>
                
                <!-- Step 5: AI System Prompt (editable, collapsed by default) -->
                <details style="margin-top: 15px;">
                    <summary style="cursor: pointer; font-weight: bold;">⚙️ Advanced: AI System Prompt (click to expand)</summary>
                    <div class="form-group" style="margin-top: 10px;">
                        <textarea id="ai-system-prompt" rows="6" style="font-family: monospace; font-size: 11px;"></textarea>
                        <button class="btn btn-secondary" onclick="resetSystemPrompt()" style="margin-top: 5px; font-size: 11px;">Reset to Default</button>
                    </div>
                </details>
                
                <!-- Step 3: Generate Amazon Content -->
                <div style="background: #cce5ff; padding: 15px; border-radius: 8px; margin-top: 20px;">
                    <h3 style="margin: 0 0 10px 0;">🚀 Step 3: Generate Amazon Content</h3>
                    <p style="font-size: 12px; color: #666; margin-bottom: 10px;">
                        Generate product images, upload to R2, and create Amazon flatfile content.
                        This data will also be used for eBay API and Etsy exports.
                    </p>
                    <button class="btn btn-success btn-lg" onclick="generateAmazonContent()" id="btn-generate-amazon" style="font-size: 16px; padding: 12px 24px;">
                        📦 Generate Amazon Content & Images
                    </button>
                    <span id="amazon-status" style="margin-left: 15px; font-size: 12px;"></span>
                </div>
                
                <!-- Progress/Output Section -->
                <div id="generation-progress" style="display: none; margin-top: 20px; background: #f8f9fa; padding: 15px; border-radius: 8px;">
                    <h4 style="margin: 0 0 10px 0;">Generation Progress</h4>
                    <div id="progress-details" style="font-size: 12px; font-family: monospace;"></div>
                </div>
                
                <!-- Generated Images Preview -->
                <div style="background: #e8f4f8; padding: 15px; border-radius: 8px; margin-top: 20px;">
                    <h3 style="margin: 0 0 10px 0;">🖼️ Generated Images Preview</h3>
                    <p style="font-size: 12px; color: #666; margin-bottom: 10px;">
                        Preview all 4 image types for each product: Main, Dimensions, Peel & Stick, Rear
                    </p>
                    <button class="btn btn-secondary" onclick="loadAllImagePreviews()" id="btn-load-previews">🔄 Load Image Previews</button>
                    <div id="all-images-preview" style="margin-top: 15px; max-height: 600px; overflow-y: auto;"></div>
                </div>
            </div>
            
            <!-- Debug Terminal -->
            <div class="card" style="margin-top: 20px;">
                <h3 style="margin-bottom: 10px;">Debug Console</h3>
                <div id="generate-debug-console" class="output-box" style="height: 200px; font-size: 11px;"></div>
            </div>
        </div>
        
        <!-- Export Tab -->
        <div id="export-panel" class="panel">
            <div class="card">
                <h2>Export Products</h2>
                <p style="margin-bottom: 15px; color: #666;">Generate lifestyle images and export to all marketplaces.</p>
                
                <!-- Step 1: Generate Lifestyle Images -->
                <div style="background: #fff3cd; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                    <h3 style="margin: 0 0 10px 0;">🎨 Step 1: Generate Lifestyle Images</h3>
                    <p style="font-size: 12px; color: #666; margin-bottom: 10px;">
                        Generate an AI background image, then overlay each product's transparent PNG to create lifestyle images.
                    </p>
                    <div style="display: flex; gap: 10px; flex-wrap: wrap; align-items: center;">
                        <button class="btn btn-warning" onclick="generateLifestyleBackground()" id="btn-lifestyle-bg">🖼️ Generate Background</button>
                        <button class="btn btn-primary" onclick="generateAllLifestyleImages()" id="btn-lifestyle-all" disabled>📸 Create Lifestyle Images</button>
                    </div>
                    <div id="lifestyle-preview" style="margin-top: 15px; display: none;">
                        <p style="font-size: 12px; font-weight: bold;">Background Preview:</p>
                        <img id="lifestyle-bg-preview" style="max-width: 300px; border-radius: 8px; border: 1px solid #ddd;">
                    </div>
                    <div id="lifestyle-images-preview" style="margin-top: 15px; display: none;">
                        <p style="font-size: 12px; font-weight: bold;">Generated Lifestyle Images:</p>
                        <div id="lifestyle-images-grid" style="display: grid; grid-template-columns: repeat(auto-fill, minmax(150px, 1fr)); gap: 10px; max-height: 300px; overflow-y: auto;"></div>
                    </div>
                    <span id="lifestyle-status" style="margin-top: 10px; display: block; font-size: 12px;"></span>
                </div>
                
                <!-- Amazon Flatfile Preview Table -->
                <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                    <h3 style="margin: 0 0 10px 0;">📋 Amazon Flatfile Preview</h3>
                    <p style="font-size: 12px; color: #666; margin-bottom: 10px;">
                        Preview of the Amazon flatfile data that will be exported. Scroll horizontally to see all columns.
                    </p>
                    <button class="btn btn-secondary" onclick="loadFlatfilePreview()" style="margin-bottom: 10px;">🔄 Refresh Preview</button>
                    <div style="overflow-x: scroll; overflow-y: auto; max-height: 450px; border: 1px solid #ddd; background: white;">
                        <table id="flatfile-preview-table" style="font-size: 10px; white-space: nowrap; border-collapse: collapse; min-width: 2000px;">
                            <thead id="flatfile-thead" style="position: sticky; top: 0; background: #343a40; color: white;">
                            </thead>
                            <tbody id="flatfile-tbody">
                                <tr><td colspan="15" style="text-align: center; color: #888; padding: 20px;">Click "Refresh Preview" to load data</td></tr>
                            </tbody>
                        </table>
                    </div>
                </div>
                
                <!-- Step 2: Export to Marketplaces -->
                <div style="background: #d4edda; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                    <h3 style="margin: 0 0 10px 0;">📦 Step 2: Export to Marketplaces</h3>
                    <p style="font-size: 12px; color: #666; margin-bottom: 10px;">
                        Generate flatfiles and publish to marketplaces. Each channel can be exported separately.
                    </p>
                    
                    <!-- Marketplace Export Buttons -->
                    <div style="display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 15px;">
                        <button class="btn btn-warning" onclick="exportAmazonFlatfile()" id="btn-export-amazon" style="padding: 10px 20px;">
                            📦 Generate Amazon Flatfile
                        </button>
                        <button class="btn btn-primary" onclick="publishToEbay()" id="btn-export-ebay" style="padding: 10px 20px;">
                            🛒 Publish to eBay (with Ads)
                        </button>
                        <button class="btn btn-danger" onclick="exportEtsyFile()" id="btn-export-etsy" style="padding: 10px 20px;">
                            🧡 Generate Etsy File
                        </button>
                    </div>
                    
                    <span id="export-all-status" style="font-size: 12px;"></span>
                    <div id="export-progress" style="display: none; margin-top: 15px; background: #f8f9fa; padding: 10px; border-radius: 4px; font-size: 12px; font-family: monospace;"></div>
                    
                    <!-- Download buttons for generated files -->
                    <div id="download-buttons" style="margin-top: 15px; display: flex; gap: 10px; flex-wrap: wrap;">
                        <button class="btn btn-outline-primary" onclick="downloadAmazonFlatfile()">📥 Download Amazon Flatfile</button>
                        <button class="btn btn-outline-warning" onclick="downloadEtsyFile()">📥 Download Etsy File</button>
                    </div>
                </div>
                
                <!-- Step 3: Download M Number Folders -->
                <div style="background: #cce5ff; padding: 15px; border-radius: 8px;">
                    <h3 style="margin: 0 0 10px 0;">📁 Step 3: Download M Number Folders</h3>
                    <p style="font-size: 12px; color: #666; margin-bottom: 10px;">
                        Generate ZIP file with full directory structure for Google Drive: 001 Design/001 MASTER FILE (SVG), 002 Images (PNG/JPEG).
                    </p>
                    <button class="btn btn-primary" onclick="downloadMNumberFolders()" id="btn-download-folders">📥 Download M Number Folders (ZIP)</button>
                    <span id="folders-status" style="margin-left: 15px; font-size: 12px;"></span>
                </div>
                
                <!-- Debug Console -->
                <div style="margin-top: 20px;">
                    <h3 style="margin-bottom: 10px;">Debug Console</h3>
                    <div id="export-debug-console" class="output-box" style="height: 150px; font-size: 11px;"></div>
                </div>
            </div>
        </div>
    </div>
    
    <!-- Add/Edit Product Modal -->
    <div id="product-modal" style="display: none; position: fixed; top: 0; left: 0; right: 0; bottom: 0; background: rgba(0,0,0,0.5); z-index: 1000;">
        <div style="background: white; max-width: 600px; margin: 50px auto; border-radius: 12px; padding: 20px;">
            <h2 id="modal-title">Add Product</h2>
            <form id="product-form">
                <div class="form-row">
                    <div class="form-group">
                        <label>M Number</label>
                        <input type="text" id="f-m_number" required placeholder="M1234">
                    </div>
                    <div class="form-group">
                        <label>EAN</label>
                        <input type="text" id="f-ean" placeholder="5056338664471">
                    </div>
                </div>
                <div class="form-group">
                    <label>Description</label>
                    <input type="text" id="f-description" required placeholder="Self Adhesive no entry aluminium sign">
                </div>
                <div class="form-row">
                    <div class="form-group">
                        <label>Size</label>
                        <select id="f-size">
                            <option value="dracula">Dracula (9.5x9.5cm) - XS</option>
                            <option value="saville">Saville (11x9.5cm) - S</option>
                            <option value="dick">Dick (14x9cm) - M</option>
                            <option value="barzan">Barzan (19x14cm) - L</option>
                            <option value="baby_jesus">Baby Jesus (29x19cm) - XL</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Color</label>
                        <select id="f-color">
                            <option value="silver">Silver</option>
                            <option value="gold">Gold</option>
                            <option value="white">White</option>
                        </select>
                    </div>
                </div>
                <div class="form-group">
                    <label>Icon Files</label>
                    <input type="text" id="f-icon_files" placeholder="no_entry.png">
                </div>
                <div class="form-row">
                    <div class="form-group">
                        <label>Orientation</label>
                        <select id="f-orientation">
                            <option value="landscape">Landscape</option>
                            <option value="portrait">Portrait</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Mounting Type</label>
                        <select id="f-mounting_type">
                            <option value="self_adhesive">Self Adhesive</option>
                            <option value="pre_drilled">Pre-Drilled</option>
                        </select>
                    </div>
                </div>
                <div style="margin-top: 20px; display: flex; gap: 10px;">
                    <button type="submit" class="btn btn-primary">Save</button>
                    <button type="button" class="btn btn-secondary" onclick="closeModal()">Cancel</button>
                </div>
            </form>
        </div>
    </div>
    
    <script>
        let products = [];
        
        function showTab(tab) {
            document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
            document.querySelector(`[onclick="showTab('${tab}')"]`).classList.add('active');
            document.getElementById(`${tab}-panel`).classList.add('active');
            
            if (tab === 'products') loadProducts();
            if (tab === 'qa') loadQAProducts();
        }
        
        async function loadProducts() {
            const resp = await fetch('/api/products');
            products = await resp.json();
            renderProductsTable();
        }
        
        function renderProductsTable() {
            const tbody = document.getElementById('products-tbody');
            tbody.innerHTML = products.map(p => `
                <tr>
                    <td><strong>${p.m_number}</strong></td>
                    <td>${p.description || ''}</td>
                    <td>${p.size || ''}</td>
                    <td>${p.color || ''}</td>
                    <td style="font-family: monospace;">${String(p.ean || '')}</td>
                    <td><span class="status-badge status-${p.qa_status || 'pending'}">${p.qa_status || 'pending'}</span></td>
                    <td>
                        <button class="btn btn-secondary" onclick="editProduct('${p.m_number}')" style="padding: 5px 10px;">Edit</button>
                        <button class="btn btn-primary" onclick="downloadProductImages('${p.m_number}')" style="padding: 5px 10px;">📦</button>
                        <button class="btn btn-danger" onclick="deleteProduct('${p.m_number}')" style="padding: 5px 10px;">Delete</button>
                    </td>
                </tr>
            `).join('');
        }
        
        // Debug console logging
        function debugLog(message, type = 'info') {
            const console = document.getElementById('debug-console');
            if (!console) return;
            const timestamp = new Date().toLocaleTimeString();
            const color = type === 'error' ? '#f00' : type === 'success' ? '#0f0' : '#0ff';
            console.innerHTML += `<span style="color: ${color}">[${timestamp}] ${message}</span>\n`;
            console.scrollTop = console.scrollHeight;
        }
        
        async function loadQAProducts() {
            debugLog('Loading products...');
            const resp = await fetch('/api/products');
            products = await resp.json();
            debugLog(`Loaded ${products.length} products`, 'success');
            renderQAGrid();
        }
        
        function renderQAGrid() {
            const grid = document.getElementById('qa-grid');
            // Group products by size and description to show variants together
            // Order: saville, dracula first (top row), then baby_jesus, dick (bottom row)
            const sizeOrder = ['saville', 'dracula', 'barzan', 'baby_jesus', 'dick'];
            const silverProducts = products
                .filter(p => p.color === 'silver')
                .sort((a, b) => sizeOrder.indexOf(a.size) - sizeOrder.indexOf(b.size));
            
            grid.innerHTML = silverProducts.map(p => {
                // Find gold and white variants
                const goldVariant = products.find(v => v.size === p.size && v.description === p.description && v.color === 'gold');
                const whiteVariant = products.find(v => v.size === p.size && v.description === p.description && v.color === 'white');
                const showOrientationToggle = p.size === 'dick' || p.size === 'baby_jesus';
                
                return `
                <div class="product-card">
                    <!-- Header with title -->
                    <div style="padding: 10px 15px; background: #667eea; color: white;">
                        <h3 style="margin: 0; font-size: 14px;">${p.description || p.m_number}</h3>
                        <span style="font-size: 11px; opacity: 0.9;">${p.size}${showOrientationToggle ? ` - ${p.orientation || 'landscape'}` : ''}</span>
                    </div>
                    
                    <!-- Large Silver Preview -->
                    <div style="padding: 15px; background: #e8e8e8; text-align: center;">
                        <img id="main-preview-${p.m_number}" src="/api/preview/${p.m_number}?t=${Date.now()}" alt="${p.m_number}" style="max-width: 100%; max-height: 250px; border-radius: 6px; box-shadow: 0 2px 8px rgba(0,0,0,0.15);">
                        <div style="margin-top: 8px; font-size: 11px;">
                            <strong>${p.m_number}</strong> - Silver
                            <button onclick="setQAStatus('${p.m_number}', 'rejected')" style="font-size: 9px; padding: 2px 6px; margin-left: 6px; cursor: pointer; background: #dc3545; color: white; border: none; border-radius: 3px;">✗ Reject</button>
                        </div>
                    </div>
                    
                    <!-- Controls Section - ABOVE variants -->
                    <div style="padding: 10px; background: #f0f0f0; border-top: 1px solid #ddd;">
                        ${showOrientationToggle ? `
                        <div style="display: flex; gap: 8px; align-items: center; margin-bottom: 8px;">
                            <label style="width: 70px; font-size: 11px;">Orientation:</label>
                            <select id="orientation-${p.m_number}" onchange="updateOrientation('${p.m_number}', this.value)" style="flex: 1; padding: 3px; font-size: 11px;">
                                <option value="landscape" ${(p.orientation || 'landscape') === 'landscape' ? 'selected' : ''}>Landscape</option>
                                <option value="portrait" ${p.orientation === 'portrait' ? 'selected' : ''}>Portrait</option>
                            </select>
                        </div>
                        ` : ''}
                        <div style="display: flex; gap: 8px; align-items: center; margin-bottom: 8px;">
                            <label style="width: 70px; font-size: 11px;">Icon Scale:</label>
                            <input type="range" id="icon-scale-${p.m_number}" min="0.5" max="1.5" step="0.01" value="${p.icon_scale || 1.0}" 
                                oninput="updateScaleDisplay('${p.m_number}', this.value)"
                                onchange="updateProductScale('${p.m_number}')"
                                style="flex: 1;">
                            <input type="number" id="icon-scale-input-${p.m_number}" min="0.5" max="1.5" step="0.01" value="${(p.icon_scale || 1.0).toFixed(2)}"
                                onchange="setIconScale('${p.m_number}', this.value)"
                                style="width: 50px; font-size: 10px; padding: 2px;">
                        </div>
                        <div style="display: flex; gap: 8px; align-items: center;">
                            <label style="width: 70px; font-size: 11px;">Position:</label>
                            <div style="display: grid; grid-template-columns: 26px 26px 26px; gap: 2px;">
                                <div></div>
                                <button type="button" onclick="moveIcon('${p.m_number}', 0, -2)" style="padding: 3px 6px; font-size: 9px; cursor: pointer;">▲</button>
                                <div></div>
                                <button type="button" onclick="moveIcon('${p.m_number}', -2, 0)" style="padding: 3px 6px; font-size: 9px; cursor: pointer;">◀</button>
                                <button type="button" onclick="centerIcon('${p.m_number}')" style="padding: 3px 5px; font-size: 7px; cursor: pointer;">⊙</button>
                                <button type="button" onclick="moveIcon('${p.m_number}', 2, 0)" style="padding: 3px 6px; font-size: 9px; cursor: pointer;">▶</button>
                                <div></div>
                                <button type="button" onclick="moveIcon('${p.m_number}', 0, 2)" style="padding: 3px 6px; font-size: 9px; cursor: pointer;">▼</button>
                                <div></div>
                            </div>
                            <span style="font-size: 9px; color: #666;" id="offset-${p.m_number}">(${(p.icon_offset_x || 0).toFixed(0)}, ${(p.icon_offset_y || 0).toFixed(0)})</span>
                            <button class="btn btn-secondary" onclick="regenerateVariants('${p.m_number}')" style="background: #6c757d; padding: 4px 8px; font-size: 10px; margin-left: auto;">🔄 Regenerate</button>
                        </div>
                    </div>
                    
                    <!-- Smaller Gold/White Variants - BELOW controls -->
                    <div style="display: flex; gap: 10px; padding: 10px; background: #f8f8f8; justify-content: center; flex-wrap: wrap;">
                        ${goldVariant ? `
                        <div style="text-align: center;">
                            <img src="/api/preview/${goldVariant.m_number}?t=${Date.now()}" alt="${goldVariant.m_number}" style="width: 80px; height: 80px; object-fit: contain; border-radius: 4px; border: 1px solid #ddd; background: white;">
                            <div style="font-size: 8px; margin-top: 3px;">
                                <strong>${goldVariant.m_number}</strong> Gold
                                <button onclick="setQAStatus('${goldVariant.m_number}', 'rejected')" style="font-size: 7px; padding: 1px 3px; margin-left: 2px; cursor: pointer; background: #dc3545; color: white; border: none; border-radius: 2px;">✗</button>
                            </div>
                        </div>
                        ` : ''}
                        ${whiteVariant ? `
                        <div style="text-align: center;">
                            <img src="/api/preview/${whiteVariant.m_number}?t=${Date.now()}" alt="${whiteVariant.m_number}" style="width: 80px; height: 80px; object-fit: contain; border-radius: 4px; border: 1px solid #ddd; background: white;">
                            <div style="font-size: 8px; margin-top: 3px;">
                                <strong>${whiteVariant.m_number}</strong> White
                                <button onclick="setQAStatus('${whiteVariant.m_number}', 'rejected')" style="font-size: 7px; padding: 1px 3px; margin-left: 2px; cursor: pointer; background: #dc3545; color: white; border: none; border-radius: 2px;">✗</button>
                            </div>
                        </div>
                        ` : ''}
                    </div>
                    
                    <!-- Action buttons -->
                    <div style="padding: 10px; background: white; border-top: 1px solid #eee;">
                        <button class="btn btn-success" onclick="approveWithVariants('${p.m_number}')" style="width: 100%;">✓ Approve All Colors</button>
                    </div>
                </div>
            `}).join('');
        }
        
        async function setQAStatus(mNumber, status) {
            await fetch(`/api/products/${mNumber}`, {
                method: 'PATCH',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({qa_status: status})
            });
            loadQAProducts();
        }
        
        function showAddProduct() {
            document.getElementById('modal-title').textContent = 'Add Product';
            document.getElementById('product-form').reset();
            document.getElementById('f-m_number').disabled = false;
            document.getElementById('product-modal').style.display = 'block';
        }
        
        function editProduct(mNumber) {
            const p = products.find(x => x.m_number === mNumber);
            if (!p) return;
            document.getElementById('modal-title').textContent = 'Edit Product';
            document.getElementById('f-m_number').value = p.m_number;
            document.getElementById('f-m_number').disabled = true;
            document.getElementById('f-description').value = p.description || '';
            document.getElementById('f-size').value = p.size || 'dracula';
            document.getElementById('f-color').value = p.color || 'silver';
            document.getElementById('f-ean').value = p.ean || '';
            document.getElementById('f-icon_files').value = p.icon_files || '';
            document.getElementById('f-orientation').value = p.orientation || 'landscape';
            document.getElementById('f-mounting_type').value = p.mounting_type || 'self_adhesive';
            document.getElementById('product-modal').style.display = 'block';
        }
        
        function closeModal() {
            document.getElementById('product-modal').style.display = 'none';
        }
        
        document.getElementById('product-form').onsubmit = async (e) => {
            e.preventDefault();
            const data = {
                m_number: document.getElementById('f-m_number').value,
                description: document.getElementById('f-description').value,
                size: document.getElementById('f-size').value,
                color: document.getElementById('f-color').value,
                ean: document.getElementById('f-ean').value,
                icon_files: document.getElementById('f-icon_files').value,
                orientation: document.getElementById('f-orientation').value,
                mounting_type: document.getElementById('f-mounting_type').value,
            };
            
            const isEdit = document.getElementById('f-m_number').disabled;
            const url = isEdit ? `/api/products/${data.m_number}` : '/api/products';
            const method = isEdit ? 'PATCH' : 'POST';
            
            await fetch(url, {
                method,
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify(data)
            });
            
            closeModal();
            loadProducts();
        };
        
        async function deleteProduct(mNumber) {
            if (!confirm(`Delete ${mNumber}?`)) return;
            await fetch(`/api/products/${mNumber}`, {method: 'DELETE'});
            loadProducts();
        }
        
        async function generateImages() {
            const output = document.getElementById('generate-output');
            output.style.display = 'block';
            output.textContent = 'Starting image generation...\\n';
            
            const resp = await fetch('/api/generate/images', {method: 'POST'});
            const reader = resp.body.getReader();
            const decoder = new TextDecoder();
            
            while (true) {
                const {done, value} = await reader.read();
                if (done) break;
                output.textContent += decoder.decode(value);
                output.scrollTop = output.scrollHeight;
            }
        }
        
        async function generateContent() {
            const theme = document.getElementById('theme').value;
            const useCases = document.getElementById('use-cases').value;
            const output = document.getElementById('generate-output');
            output.style.display = 'block';
            output.textContent = 'Starting content generation...\\n';
            
            const resp = await fetch(`/api/generate/content?theme=${encodeURIComponent(theme)}&use_cases=${encodeURIComponent(useCases)}`, {method: 'POST'});
            const reader = resp.body.getReader();
            const decoder = new TextDecoder();
            
            while (true) {
                const {done, value} = await reader.read();
                if (done) break;
                output.textContent += decoder.decode(value);
                output.scrollTop = output.scrollHeight;
            }
        }
        
        async function runFullPipeline() {
            const theme = document.getElementById('theme').value;
            const useCases = document.getElementById('use-cases').value;
            const output = document.getElementById('generate-output');
            output.style.display = 'block';
            output.textContent = 'Starting full pipeline...\\n';
            
            const resp = await fetch(`/api/generate/full?theme=${encodeURIComponent(theme)}&use_cases=${encodeURIComponent(useCases)}`, {method: 'POST'});
            const reader = resp.body.getReader();
            const decoder = new TextDecoder();
            
            while (true) {
                const {done, value} = await reader.read();
                if (done) break;
                output.textContent += decoder.decode(value);
                output.scrollTop = output.scrollHeight;
            }
        }
        
        async function exportFlatfile() {
            const status = document.getElementById('export-status');
            status.innerHTML = '<div class="alert alert-info">Generating Amazon flatfile...</div>';
            
            try {
                const resp = await fetch('/api/export/flatfile', {method: 'POST'});
                if (resp.ok) {
                    const blob = await resp.blob();
                    const url = URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `amazon_flatfile_${new Date().toISOString().slice(0,10)}.xlsx`;
                    a.click();
                    status.innerHTML = '<div class="alert alert-success">Amazon flatfile downloaded!</div>';
                } else {
                    status.innerHTML = '<div class="alert alert-error">Failed to generate flatfile</div>';
                }
            } catch (e) {
                status.innerHTML = `<div class="alert alert-error">Error: ${e.message}</div>`;
            }
        }
        
        async function exportEbay() {
            const status = document.getElementById('export-status');
            status.innerHTML = '<div class="alert alert-info">Generating eBay CSV...</div>';
            
            try {
                const resp = await fetch('/api/export/ebay', {method: 'POST'});
                if (resp.ok) {
                    const blob = await resp.blob();
                    const url = URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `ebay_listings_${new Date().toISOString().slice(0,10)}.csv`;
                    a.click();
                    status.innerHTML = '<div class="alert alert-success">eBay CSV downloaded!</div>';
                } else {
                    status.innerHTML = '<div class="alert alert-error">Failed to generate eBay CSV</div>';
                }
            } catch (e) {
                status.innerHTML = `<div class="alert alert-error">Error: ${e.message}</div>`;
            }
        }
        
        async function exportEtsy() {
            const status = document.getElementById('export-status');
            status.innerHTML = '<div class="alert alert-info">Generating Etsy Shop Uploader...</div>';
            
            try {
                const resp = await fetch('/api/export/etsy', {method: 'POST'});
                if (resp.ok) {
                    const blob = await resp.blob();
                    const url = URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `etsy_shop_uploader_${new Date().toISOString().slice(0,10)}.xlsx`;
                    a.click();
                    status.innerHTML = '<div class="alert alert-success">Etsy Shop Uploader downloaded!</div>';
                } else {
                    status.innerHTML = '<div class="alert alert-error">Failed to generate Etsy file</div>';
                }
            } catch (e) {
                status.innerHTML = `<div class="alert alert-error">Error: ${e.message}</div>`;
            }
        }
        
        async function exportAllImages() {
            const status = document.getElementById('export-status');
            status.innerHTML = '<div class="alert alert-info">Generating images ZIP... This may take a while.</div>';
            
            try {
                const resp = await fetch('/api/export/images', {method: 'POST'});
                if (resp.ok) {
                    const blob = await resp.blob();
                    const url = URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `product_images_${new Date().toISOString().slice(0,10)}.zip`;
                    a.click();
                    status.innerHTML = '<div class="alert alert-success">Images ZIP downloaded!</div>';
                } else {
                    status.innerHTML = '<div class="alert alert-error">Failed to generate images ZIP</div>';
                }
            } catch (e) {
                status.innerHTML = `<div class="alert alert-error">Error: ${e.message}</div>`;
            }
        }
        
        async function downloadProductImages(mNumber) {
            window.location.href = `/api/export/images/${mNumber}`;
        }
        
        async function publishToEbay(withAds) {
            const status = document.getElementById('export-status');
            status.innerHTML = '<div class="alert alert-info">Publishing to eBay... This may take a minute.</div>';
            
            try {
                const resp = await fetch(`/api/ebay/publish?promote=${withAds}`, {method: 'POST'});
                const data = await resp.json();
                
                if (data.success) {
                    let msg = `eBay listing created! ${data.products} products published.`;
                    if (data.url) {
                        msg += ` <a href="${data.url}" target="_blank">View on eBay</a>`;
                    }
                    if (data.promoted) {
                        msg += ' (Auto-promoted with 5% ad rate)';
                    }
                    status.innerHTML = `<div class="alert alert-success">${msg}</div>`;
                } else {
                    status.innerHTML = `<div class="alert alert-error">Error: ${data.error}</div>`;
                }
            } catch (e) {
                status.innerHTML = `<div class="alert alert-error">Error: ${e.message}</div>`;
            }
        }
        
        async function exportMNumberFolders() {
            const status = document.getElementById('export-status');
            status.innerHTML = '<div class="alert alert-info">Generating M Number folders... This may take a while.</div>';
            
            try {
                const resp = await fetch('/api/export/m-number-folders', {method: 'POST'});
                if (resp.ok) {
                    const blob = await resp.blob();
                    const url = URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `m_number_folders_${new Date().toISOString().slice(0,10)}.zip`;
                    a.click();
                    status.innerHTML = '<div class="alert alert-success">M Number folders downloaded!</div>';
                } else {
                    status.innerHTML = '<div class="alert alert-error">Failed to generate M Number folders</div>';
                }
            } catch (e) {
                status.innerHTML = `<div class="alert alert-error">Error: ${e.message}</div>`;
            }
        }
        
        async function generateLifestyleImages() {
            const status = document.getElementById('export-status');
            status.innerHTML = '<div class="alert alert-info">Generating lifestyle images with DALL-E 3... This may take a few minutes.</div>';
            
            try {
                const resp = await fetch('/api/generate/lifestyle', {method: 'POST'});
                const data = await resp.json();
                
                if (data.success) {
                    status.innerHTML = `<div class="alert alert-success">Generated ${data.generated} lifestyle images! (${data.skipped} skipped, ${data.failed} failed)</div>`;
                } else {
                    status.innerHTML = `<div class="alert alert-error">Error: ${data.error}</div>`;
                }
            } catch (e) {
                status.innerHTML = `<div class="alert alert-error">Error: ${e.message}</div>`;
            }
        }
        
        // Debounce timers for updates
        const updateTimers = {};
        
        function updateScaleDisplay(mNumber, value) {
            document.getElementById(`icon-scale-input-${mNumber}`).value = parseFloat(value).toFixed(2);
        }
        
        function setIconScale(mNumber, value) {
            const scale = parseFloat(value);
            if (scale >= 0.5 && scale <= 1.5) {
                document.getElementById(`icon-scale-${mNumber}`).value = scale;
                updateProductScale(mNumber);
            }
        }
        
        function updateProductScale(mNumber) {
            const iconScale = document.getElementById(`icon-scale-${mNumber}`).value;
            
            // Debounce: wait 300ms after last change before updating
            if (updateTimers[mNumber]) {
                clearTimeout(updateTimers[mNumber]);
            }
            updateTimers[mNumber] = setTimeout(async () => {
                try {
                    const resp = await fetch(`/api/products/${mNumber}/scale`, {
                        method: 'PATCH',
                        headers: {'Content-Type': 'application/json'},
                        body: JSON.stringify({icon_scale: parseFloat(iconScale)})
                    });
                    if (resp.ok) {
                        refreshProductImage(mNumber);
                        const p = products.find(x => x.m_number === mNumber);
                        if (p) {
                            p.icon_scale = parseFloat(iconScale);
                        }
                    }
                } catch (e) {
                    console.error('Failed to update scale:', e);
                }
            }, 300);
        }
        
        async function approveWithVariants(mNumber) {
            // Get the silver product's settings
            const silverProduct = products.find(x => x.m_number === mNumber);
            if (!silverProduct) return;
            
            const { icon_scale, icon_offset_x, icon_offset_y, size, description } = silverProduct;
            
            // Find gold and white variants (same size and description)
            const variants = products.filter(p => 
                p.size === size && 
                p.description === description && 
                (p.color === 'gold' || p.color === 'white')
            );
            
            // Apply settings to all variants and approve
            try {
                // Approve silver
                await fetch(`/api/products/${mNumber}`, {
                    method: 'PATCH',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({qa_status: 'approved'})
                });
                
                // Apply settings to gold/white variants and approve
                for (const variant of variants) {
                    await fetch(`/api/products/${variant.m_number}/scale`, {
                        method: 'PATCH',
                        headers: {'Content-Type': 'application/json'},
                        body: JSON.stringify({icon_scale: icon_scale})
                    });
                    await fetch(`/api/products/${variant.m_number}/position`, {
                        method: 'PATCH',
                        headers: {'Content-Type': 'application/json'},
                        body: JSON.stringify({icon_offset_x: icon_offset_x || 0, icon_offset_y: icon_offset_y || 0})
                    });
                    await fetch(`/api/products/${variant.m_number}`, {
                        method: 'PATCH',
                        headers: {'Content-Type': 'application/json'},
                        body: JSON.stringify({qa_status: 'approved'})
                    });
                }
                
                loadQAProducts();
            } catch (e) {
                console.error('Failed to approve variants:', e);
            }
        }
        
        async function regenerateVariants(mNumber) {
            // Get the silver product's settings
            const silverProduct = products.find(x => x.m_number === mNumber);
            if (!silverProduct) return;
            
            // Show loading state
            const btn = event.target;
            const originalText = btn.innerHTML;
            btn.innerHTML = '⏳ Working...';
            btn.disabled = true;
            debugLog(`Regenerating variants for ${mNumber}...`);
            
            const { icon_scale, icon_offset_x, icon_offset_y, size, description } = silverProduct;
            
            // Find gold and white variants
            const variants = products.filter(p => 
                p.size === size && 
                p.description === description && 
                (p.color === 'gold' || p.color === 'white')
            );
            
            // Apply silver settings to all variants and refresh images
            try {
                for (const variant of variants) {
                    debugLog(`  Updating ${variant.m_number}...`);
                    await fetch(`/api/products/${variant.m_number}/scale`, {
                        method: 'PATCH',
                        headers: {'Content-Type': 'application/json'},
                        body: JSON.stringify({icon_scale: icon_scale || 1.0})
                    });
                    await fetch(`/api/products/${variant.m_number}/position`, {
                        method: 'PATCH',
                        headers: {'Content-Type': 'application/json'},
                        body: JSON.stringify({icon_offset_x: icon_offset_x || 0, icon_offset_y: icon_offset_y || 0})
                    });
                }
                
                debugLog(`Regenerated ${variants.length + 1} variants`, 'success');
                
                // Refresh all variant images
                refreshProductImage(mNumber);
                for (const variant of variants) {
                    refreshProductImage(variant.m_number);
                }
            } catch (e) {
                debugLog(`Error: ${e.message}`, 'error');
                console.error('Failed to regenerate variants:', e);
            } finally {
                btn.innerHTML = originalText;
                btn.disabled = false;
            }
        }
        
        async function moveIcon(mNumber, dx, dy) {
            const p = products.find(x => x.m_number === mNumber);
            if (!p) return;
            
            const newX = (p.icon_offset_x || 0) + dx;
            const newY = (p.icon_offset_y || 0) + dy;
            
            try {
                const resp = await fetch(`/api/products/${mNumber}/position`, {
                    method: 'PATCH',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({icon_offset_x: newX, icon_offset_y: newY})
                });
                if (resp.ok) {
                    p.icon_offset_x = newX;
                    p.icon_offset_y = newY;
                    document.getElementById(`offset-${mNumber}`).textContent = `(${newX.toFixed(0)}, ${newY.toFixed(0)})`;
                    refreshProductImage(mNumber);
                }
            } catch (e) {
                console.error('Failed to move icon:', e);
            }
        }
        
        async function centerIcon(mNumber) {
            try {
                const resp = await fetch(`/api/products/${mNumber}/position`, {
                    method: 'PATCH',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({icon_offset_x: 0, icon_offset_y: 0})
                });
                if (resp.ok) {
                    const p = products.find(x => x.m_number === mNumber);
                    if (p) {
                        p.icon_offset_x = 0;
                        p.icon_offset_y = 0;
                    }
                    document.getElementById(`offset-${mNumber}`).textContent = '(0, 0)';
                    refreshProductImage(mNumber);
                }
            } catch (e) {
                console.error('Failed to center icon:', e);
            }
        }
        
        function refreshProductImage(mNumber) {
            const img = document.querySelector(`img[src*="/api/preview/${mNumber}"]`);
            if (img) {
                img.src = `/api/preview/${mNumber}?t=${Date.now()}`;
            }
        }
        
        async function updateOrientation(mNumber, orientation) {
            const silverProduct = products.find(x => x.m_number === mNumber);
            if (!silverProduct) return;
            
            const { size, description } = silverProduct;
            
            // Find all color variants
            const allVariants = products.filter(p => 
                p.size === size && p.description === description
            );
            
            try {
                // Update orientation for all variants
                for (const variant of allVariants) {
                    await fetch(`/api/products/${variant.m_number}`, {
                        method: 'PATCH',
                        headers: {'Content-Type': 'application/json'},
                        body: JSON.stringify({orientation: orientation})
                    });
                    variant.orientation = orientation;
                }
                
                // Refresh all images
                loadQAProducts();
            } catch (e) {
                console.error('Failed to update orientation:', e);
            }
        }
        
        // ============ GENERATE TAB FUNCTIONS ============
        
        // Progress indicator management
        let progressInterval = null;
        let progressDots = 0;
        let progressStartTime = null;
        
        function startProgress(message) {
            stopProgress();
            progressStartTime = Date.now();
            progressDots = 0;
            const console = document.getElementById('generate-debug-console');
            if (!console) return;
            
            // Add initial message with progress span
            const timestamp = new Date().toLocaleTimeString();
            console.innerHTML += `<span style="color: #ff0" id="progress-line">[${timestamp}] ⏳ ${message} <span id="progress-dots"></span> <span id="progress-time"></span></span>\n`;
            console.scrollTop = console.scrollHeight;
            
            // Update dots and elapsed time every 500ms
            progressInterval = setInterval(() => {
                progressDots = (progressDots + 1) % 4;
                const dots = '.'.repeat(progressDots + 1);
                const elapsed = Math.floor((Date.now() - progressStartTime) / 1000);
                const dotsSpan = document.getElementById('progress-dots');
                const timeSpan = document.getElementById('progress-time');
                if (dotsSpan) dotsSpan.textContent = dots;
                if (timeSpan) timeSpan.textContent = `(${elapsed}s)`;
            }, 500);
        }
        
        function stopProgress() {
            if (progressInterval) {
                clearInterval(progressInterval);
                progressInterval = null;
            }
            // Remove the progress line
            const progressLine = document.getElementById('progress-line');
            if (progressLine) progressLine.remove();
        }
        
        // Debug logging for Generate tab
        function genLog(message, type = 'info') {
            stopProgress(); // Stop any running progress indicator
            const console = document.getElementById('generate-debug-console');
            if (!console) return;
            const timestamp = new Date().toLocaleTimeString();
            const color = type === 'error' ? '#f00' : type === 'success' ? '#0f0' : '#0ff';
            console.innerHTML += `<span style="color: ${color}">[${timestamp}] ${message}</span>\n`;
            console.scrollTop = console.scrollHeight;
        }
        
        // Default system prompt (no internal part names)
        const DEFAULT_SYSTEM_PROMPT = `You are an expert product content writer for Amazon marketplace listings.
You will be provided with:
1. A summary of the product range (sizes, shapes, colors)
2. Sample images showing the actual products
3. Theme and use case information

IMPORTANT: The products come in MULTIPLE sizes and shapes:
- Small rectangular: 115x95mm
- Medium rectangular: 140x90mm
- Large rectangular: 194x143mm
- Extra large rectangular: 290x190mm
- Circular: 95mm diameter
- All signs come in 3 colors: silver, gold, and white

Generate content that accurately describes ALL variants, not just one shape.
Include dimensions in product descriptions.
Write compelling Amazon-style titles and bullet points.`;
        
        // Analyze products with AI to auto-populate theme and use cases
        async function analyzeProducts() {
            const btn = document.getElementById('btn-analyze');
            const status = document.getElementById('analyze-status');
            btn.disabled = true;
            btn.innerHTML = '⏳ Analyzing...';
            status.textContent = '';
            genLog('Starting product analysis with AI...');
            
            try {
                // Get sample image M numbers
                const sampleMNumbers = Array.from(document.querySelectorAll('#sample-images img'))
                    .map(img => {
                        const match = img.src.match(/\/api\/preview\/([^?]+)/);
                        return match ? match[1] : null;
                    })
                    .filter(Boolean);
                
                genLog(`Sending ${sampleMNumbers.length} images for analysis...`);
                startProgress('Waiting for AI response');
                
                // Add 2 minute timeout for AI analysis
                const controller = new AbortController();
                const timeout = setTimeout(() => controller.abort(), 120000);
                
                const resp = await fetch('/api/analyze/products', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({ sample_m_numbers: sampleMNumbers }),
                    signal: controller.signal
                });
                clearTimeout(timeout);
                
                const data = await resp.json();
                
                if (data.success) {
                    document.getElementById('theme').value = data.theme || '';
                    document.getElementById('use-cases').value = data.use_cases || '';
                    status.textContent = '✓ Fields populated!';
                    status.style.color = 'green';
                    genLog('Analysis complete - fields populated', 'success');
                } else {
                    status.textContent = `Error: ${data.error}`;
                    status.style.color = 'red';
                    genLog(`Analysis failed: ${data.error}`, 'error');
                }
            } catch (e) {
                if (e.name === 'AbortError') {
                    status.textContent = 'Request timed out - try again';
                    status.style.color = 'red';
                    genLog('Analysis timed out after 2 minutes', 'error');
                } else {
                    status.textContent = 'Error analyzing products';
                    status.style.color = 'red';
                    genLog(`Analysis error: ${e.message}`, 'error');
                }
            } finally {
                stopProgress();
                btn.disabled = false;
                btn.innerHTML = '🔍 Analyze Products with AI';
            }
        }

        function resetSystemPrompt() {
            document.getElementById('ai-system-prompt').value = DEFAULT_SYSTEM_PROMPT;
            genLog('System prompt reset to default');
        }
        
        // Load product summary and sample images when Generate tab is shown
        function loadGenerateTabData() {
            genLog('Loading product summary...');
            
            // Build product summary
            const sizes = {};
            const colors = new Set();
            const descriptions = new Set();
            
            products.forEach(p => {
                const size = p.size || 'unknown';
                const color = p.color || 'unknown';
                sizes[size] = (sizes[size] || 0) + 1;
                colors.add(color);
                if (p.description) descriptions.add(p.description);
            });
            
            // Size dimensions (display names without internal part names)
            const sizeDims = {
                'saville': 'Small (115x95mm, rectangular)',
                'dick': 'Medium (140x90mm, rectangular)',
                'barzan': 'Large (194x143mm, rectangular)',
                'dracula': 'Circular (95mm diameter)',
                'baby_jesus': 'Extra Large (290x190mm, rectangular)'
            };
            
            let summaryHTML = `<strong>Total Products:</strong> ${products.length}<br>`;
            summaryHTML += `<strong>Colors:</strong> ${Array.from(colors).join(', ')}<br>`;
            summaryHTML += `<strong>Sizes:</strong><br>`;
            for (const [size, count] of Object.entries(sizes)) {
                const dims = sizeDims[size] || 'unknown dimensions';
                summaryHTML += `&nbsp;&nbsp;• ${dims}: ${count} products<br>`;
            }
            summaryHTML += `<strong>Product Types:</strong> ${descriptions.size} unique designs`;
            
            document.getElementById('product-summary').innerHTML = summaryHTML;
            genLog(`Found ${products.length} products across ${Object.keys(sizes).length} sizes`, 'success');
            
            // Load sample images - one of each size (silver only for variety)
            const sampleImages = [];
            const seenSizes = new Set();
            
            for (const p of products) {
                if (p.color === 'silver' && !seenSizes.has(p.size)) {
                    sampleImages.push(p);
                    seenSizes.add(p.size);
                }
                if (sampleImages.length >= 5) break;
            }
            
            let imagesHTML = '';
            for (const p of sampleImages) {
                imagesHTML += `
                    <div style="text-align: center;">
                        <img src="/api/preview/${p.m_number}?t=${Date.now()}" 
                             style="width: 100px; height: 80px; object-fit: contain; border: 1px solid #ddd; border-radius: 4px; background: white;">
                        <div style="font-size: 10px; color: #666;">${p.size}</div>
                    </div>
                `;
            }
            
            if (imagesHTML) {
                document.getElementById('sample-images').innerHTML = imagesHTML;
                genLog(`Loaded ${sampleImages.length} sample images for AI context`, 'success');
            } else {
                document.getElementById('sample-images').innerHTML = '<em style="color: #888;">No products available</em>';
            }
            
            // Set default system prompt if empty
            const promptField = document.getElementById('ai-system-prompt');
            if (!promptField.value) {
                promptField.value = DEFAULT_SYSTEM_PROMPT;
            }
        }
        
        function previewAIPrompt() {
            const theme = document.getElementById('theme').value || '(not specified)';
            const useCases = document.getElementById('use-cases').value || '(not specified)';
            const systemPrompt = document.getElementById('ai-system-prompt').value;
            const summary = document.getElementById('product-summary').innerText;
            
            const fullPrompt = `=== SYSTEM PROMPT ===
${systemPrompt}

=== PRODUCT SUMMARY ===
${summary}

=== USER INPUT ===
Theme: ${theme}
Use Cases: ${useCases}

=== SAMPLE IMAGES ===
[${document.querySelectorAll('#sample-images img').length} images will be attached]`;
            
            alert(fullPrompt);
            genLog('Previewed full AI prompt');
        }
        
        async function generateContent() {
            genLog('Starting content generation...');
            const btn = event.target;
            btn.disabled = true;
            btn.innerHTML = '⏳ Generating...';
            
            try {
                const theme = document.getElementById('theme').value;
                const useCases = document.getElementById('use-cases').value;
                const systemPrompt = document.getElementById('ai-system-prompt').value;
                
                // Get sample image M numbers
                const sampleMNumbers = Array.from(document.querySelectorAll('#sample-images img'))
                    .map(img => {
                        const match = img.src.match(/\/api\/preview\/([^?]+)/);
                        return match ? match[1] : null;
                    })
                    .filter(Boolean);
                
                genLog(`Sending ${sampleMNumbers.length} sample images to AI...`);
                
                const resp = await fetch('/api/generate/content', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        theme,
                        use_cases: useCases,
                        system_prompt: systemPrompt,
                        sample_m_numbers: sampleMNumbers
                    })
                });
                
                const data = await resp.json();
                
                if (data.success) {
                    genLog('Content generated successfully!', 'success');
                    document.getElementById('ai-response').value = data.content;
                    document.getElementById('ai-response-section').style.display = 'block';
                } else {
                    genLog(`Error: ${data.error}`, 'error');
                    alert('Error: ' + data.error);
                }
            } catch (e) {
                genLog(`Error: ${e.message}`, 'error');
                console.error('Failed to generate content:', e);
            } finally {
                btn.disabled = false;
                btn.innerHTML = '📝 Generate Content';
            }
        }
        
        async function generateAmazonContent() {
            const btn = document.getElementById('btn-generate-amazon');
            const status = document.getElementById('amazon-status');
            const progressDiv = document.getElementById('generation-progress');
            const progressDetails = document.getElementById('progress-details');
            
            btn.disabled = true;
            btn.innerHTML = '⏳ Generating...';
            status.textContent = '';
            progressDiv.style.display = 'block';
            progressDetails.innerHTML = '';
            
            const theme = document.getElementById('theme').value;
            const useCases = document.getElementById('use-cases').value;
            const systemPrompt = document.getElementById('ai-system-prompt').value;
            
            // Get sample image M numbers
            const sampleMNumbers = Array.from(document.querySelectorAll('#sample-images img'))
                .map(img => {
                    const match = img.src.match(/\/api\/preview\/([^?]+)/);
                    return match ? match[1] : null;
                })
                .filter(Boolean);
            
            try {
                // Step 1: Generate images and upload to R2
                genLog('Step 1: Generating product images...');
                progressDetails.innerHTML += '📸 Generating product images...<br>';
                startProgress('Generating images');
                
                const imgResp = await fetch('/api/generate/images', {method: 'POST'});
                const imgData = await imgResp.json();
                
                if (imgData.success) {
                    genLog(`Images generated for ${imgData.count} products`, 'success');
                    progressDetails.innerHTML += `✅ Generated images for ${imgData.count} products<br>`;
                } else {
                    throw new Error(imgData.error || 'Image generation failed');
                }
                
                // Step 2: Generate AI content (with 2 minute timeout)
                genLog('Step 2: Generating AI content (may take up to 2 minutes)...');
                progressDetails.innerHTML += '🤖 Generating AI content (may take up to 2 minutes)...<br>';
                startProgress('Waiting for AI response');
                
                const contentController = new AbortController();
                const contentTimeout = setTimeout(() => contentController.abort(), 120000);
                
                const contentResp = await fetch('/api/generate/content', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        theme,
                        use_cases: useCases,
                        system_prompt: systemPrompt,
                        sample_m_numbers: sampleMNumbers
                    }),
                    signal: contentController.signal
                });
                clearTimeout(contentTimeout);
                
                const contentData = await contentResp.json();
                
                if (contentData.success) {
                    genLog('AI content generated', 'success');
                    progressDetails.innerHTML += '✅ AI content generated<br>';
                } else {
                    throw new Error(contentData.error || 'Content generation failed');
                }
                
                // Step 3: Generate Amazon flatfile
                genLog('Step 3: Generating Amazon flatfile...');
                progressDetails.innerHTML += '📄 Generating Amazon flatfile...<br>';
                startProgress('Preparing flatfile');
                
                const flatfileResp = await fetch('/api/generate/amazon-flatfile', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        theme,
                        use_cases: useCases,
                        ai_content: contentData.content
                    })
                });
                
                const flatfileData = await flatfileResp.json();
                
                if (flatfileData.success) {
                    genLog('Amazon flatfile generated!', 'success');
                    progressDetails.innerHTML += '✅ Amazon flatfile ready<br>';
                    progressDetails.innerHTML += `<br><strong>Complete!</strong> ${flatfileData.product_count} products processed.<br>`;
                    status.textContent = '✓ Complete!';
                    status.style.color = 'green';
                } else {
                    throw new Error(flatfileData.error || 'Flatfile generation failed');
                }
                
            } catch (e) {
                if (e.name === 'AbortError') {
                    genLog('Request timed out after 2 minutes - try again', 'error');
                    progressDetails.innerHTML += `❌ Request timed out - AI may be slow, try again<br>`;
                    status.textContent = 'Timeout - try again';
                } else {
                    genLog(`Error: ${e.message}`, 'error');
                    progressDetails.innerHTML += `❌ Error: ${e.message}<br>`;
                    status.textContent = 'Error - see console';
                }
                status.style.color = 'red';
            } finally {
                stopProgress();
                btn.disabled = false;
                btn.innerHTML = '📦 Generate Amazon Content & Images';
            }
        }
        
        // Override showTab to load Generate tab data
        const originalShowTab = showTab;
        showTab = function(tabName) {
            originalShowTab(tabName);
            if (tabName === 'generate') {
                loadGenerateTabData();
            }
        };
        
        // Load all image previews for Generate tab
        async function loadAllImagePreviews() {
            const btn = document.getElementById('btn-load-previews');
            const container = document.getElementById('all-images-preview');
            btn.disabled = true;
            btn.innerHTML = '⏳ Loading...';
            container.innerHTML = '<p style="color: #888;">Loading image previews...</p>';
            
            try {
                const resp = await fetch('/api/products');
                const products = await resp.json();
                
                if (!products || products.length === 0) {
                    container.innerHTML = '<p style="color: #888;">No products found.</p>';
                    return;
                }
                
                const imageTypes = ['main', 'dimensions', 'peel_and_stick', 'rear'];
                const typeLabels = {
                    'main': '1. Main',
                    'dimensions': '2. Dimensions', 
                    'peel_and_stick': '3. Peel & Stick',
                    'rear': '4. Rear'
                };
                
                let html = '';
                for (const product of products) {
                    const mNumber = product.m_number;
                    html += `
                        <div style="border: 1px solid #ddd; border-radius: 8px; padding: 10px; margin-bottom: 15px; background: white;">
                            <h4 style="margin: 0 0 10px 0; font-size: 14px;">${mNumber} - ${product.description || 'No description'}</h4>
                            <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px;">
                    `;
                    
                    for (const imgType of imageTypes) {
                        const imgUrl = '/api/export/images/' + mNumber + '?type=' + imgType + '&t=' + Date.now();
                        html += `
                            <div style="text-align: center;">
                                <p style="font-size: 11px; margin: 0 0 5px 0; font-weight: bold;">${typeLabels[imgType]}</p>
                                <img src="${imgUrl}" 
                                     style="max-width: 100%; max-height: 150px; border: 1px solid #eee; border-radius: 4px;"
                                     onerror="this.style.display='none'; this.nextElementSibling.style.display='block';"
                                     loading="lazy">
                                <p style="display: none; font-size: 10px; color: #999;">Not available</p>
                            </div>
                        `;
                    }
                    
                    html += `
                            </div>
                        </div>
                    `;
                }
                
                container.innerHTML = html;
                genLog('Loaded image previews for ' + products.length + ' products', 'success');
            } catch (e) {
                container.innerHTML = '<p style="color: red;">Error loading previews: ' + e.message + '</p>';
                genLog('Error loading previews: ' + e.message, 'error');
            } finally {
                btn.disabled = false;
                btn.innerHTML = '🔄 Load Image Previews';
            }
        }
        
        // ============ EXPORT TAB FUNCTIONS ============
        
        let lifestyleBackgroundUrl = null;
        
        function exportLog(message, type = 'info') {
            const console = document.getElementById('export-debug-console');
            if (!console) return;
            const timestamp = new Date().toLocaleTimeString();
            const color = type === 'error' ? '#f00' : type === 'success' ? '#0f0' : '#0ff';
            console.innerHTML += `<span style="color: ${color}">[${timestamp}] ${message}</span>\n`;
            console.scrollTop = console.scrollHeight;
        }
        
        // Step 1: Generate lifestyle background with DALL-E
        async function generateLifestyleBackground() {
            const btn = document.getElementById('btn-lifestyle-bg');
            const status = document.getElementById('lifestyle-status');
            btn.disabled = true;
            btn.innerHTML = '⏳ Generating...';
            status.textContent = '';
            exportLog('Generating lifestyle background with DALL-E...');
            
            try {
                const resp = await fetch('/api/export/lifestyle-background', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({})
                });
                const data = await resp.json();
                
                if (data.success) {
                    lifestyleBackgroundUrl = data.background_url;
                    document.getElementById('lifestyle-bg-preview').src = data.background_url;
                    document.getElementById('lifestyle-preview').style.display = 'block';
                    document.getElementById('btn-lifestyle-all').disabled = false;
                    status.textContent = '✓ Background generated!';
                    status.style.color = 'green';
                    exportLog('Background generated successfully', 'success');
                } else {
                    status.textContent = `Error: ${data.error}`;
                    status.style.color = 'red';
                    exportLog(`Error: ${data.error}`, 'error');
                }
            } catch (e) {
                status.textContent = 'Error generating background';
                status.style.color = 'red';
                exportLog(`Error: ${e.message}`, 'error');
            } finally {
                btn.disabled = false;
                btn.innerHTML = '🖼️ Generate Background';
            }
        }
        
        // Step 1b: Create lifestyle images for all products
        async function generateAllLifestyleImages() {
            const btn = document.getElementById('btn-lifestyle-all');
            const status = document.getElementById('lifestyle-status');
            btn.disabled = true;
            btn.innerHTML = '⏳ Creating...';
            exportLog('Creating lifestyle images for all products...');
            
            try {
                const resp = await fetch('/api/export/lifestyle-images', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({ background_url: lifestyleBackgroundUrl })
                });
                const data = await resp.json();
                
                if (data.success) {
                    status.textContent = `✓ Created ${data.count} lifestyle images and uploaded to R2!`;
                    status.style.color = 'green';
                    exportLog(`Created ${data.count} lifestyle images`, 'success');
                    
                    // Show preview grid of lifestyle images
                    if (data.images && data.images.length > 0) {
                        const previewDiv = document.getElementById('lifestyle-images-preview');
                        const grid = document.getElementById('lifestyle-images-grid');
                        previewDiv.style.display = 'block';
                        grid.innerHTML = data.images.map(img => `
                            <div style="text-align: center;">
                                <img src="${img.url}" style="width: 140px; height: 100px; object-fit: contain; border: 1px solid #ddd; border-radius: 4px; background: white;">
                                <div style="font-size: 9px; color: #666;">${img.m_number}</div>
                            </div>
                        `).join('');
                        exportLog(`Showing ${data.images.length} lifestyle image previews`, 'success');
                    }
                } else {
                    status.textContent = `Error: ${data.error}`;
                    status.style.color = 'red';
                    exportLog(`Error: ${data.error}`, 'error');
                }
            } catch (e) {
                status.textContent = 'Error creating lifestyle images';
                status.style.color = 'red';
                exportLog(`Error: ${e.message}`, 'error');
            } finally {
                btn.disabled = false;
                btn.innerHTML = '📸 Create Lifestyle Images';
            }
        }
        
        // Separate marketplace export functions
        async function exportAmazonFlatfile() {
            const btn = document.getElementById('btn-export-amazon');
            const status = document.getElementById('export-all-status');
            btn.disabled = true;
            btn.innerHTML = '⏳ Generating...';
            exportLog('Generating Amazon flatfile...');
            
            try {
                const theme = document.getElementById('theme')?.value || '';
                const useCases = document.getElementById('use-cases')?.value || '';
                
                const resp = await fetch('/api/export/amazon', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({ theme, use_cases: useCases })
                });
                
                if (!resp.ok) {
                    const errorData = await resp.json();
                    throw new Error(errorData.error || 'Failed to generate flatfile');
                }
                
                // Download the file
                const blob = await resp.blob();
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = `amazon_flatfile_${new Date().toISOString().slice(0,10).replace(/-/g,'')}.xlsx`;
                document.body.appendChild(a);
                a.click();
                document.body.removeChild(a);
                window.URL.revokeObjectURL(url);
                
                status.textContent = '✓ Amazon flatfile downloaded!';
                status.style.color = 'green';
                exportLog('Amazon flatfile downloaded successfully', 'success');
            } catch (e) {
                status.textContent = 'Error generating Amazon flatfile';
                status.style.color = 'red';
                exportLog(`Error: ${e.message}`, 'error');
            } finally {
                btn.disabled = false;
                btn.innerHTML = '📦 Generate Amazon Flatfile';
            }
        }
        
        async function publishToEbay() {
            const btn = document.getElementById('btn-export-ebay');
            const status = document.getElementById('export-all-status');
            const progress = document.getElementById('export-progress');
            btn.disabled = true;
            btn.innerHTML = '⏳ Publishing...';
            progress.style.display = 'block';
            progress.innerHTML = '🛒 Connecting to eBay API...<br>';
            exportLog('Publishing to eBay (this may take up to 2 minutes)...');
            exportLog('eBay: Connecting to eBay API...');
            
            try {
                const ebayController = new AbortController();
                const ebayTimeout = setTimeout(() => ebayController.abort(), 120000);
                
                exportLog('eBay: Sending listing request with ads enabled...');
                const resp = await fetch('/api/ebay/publish', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({ with_ads: true }),
                    signal: ebayController.signal
                });
                clearTimeout(ebayTimeout);
                
                exportLog(`eBay: Response received (status ${resp.status})`);
                const data = await resp.json();
                
                if (data.success) {
                    progress.innerHTML += `✅ eBay: ${data.count || 0} listings published<br>`;
                    status.textContent = '✓ eBay listings published!';
                    status.style.color = 'green';
                    exportLog(`eBay: Listing ID: ${data.listing_id}`, 'success');
                    if (data.url) exportLog(`eBay: URL: ${data.url}`, 'success');
                    if (data.promoted) exportLog('eBay: Ads enabled at 5% rate', 'success');
                    exportLog('eBay listings published', 'success');
                } else {
                    progress.innerHTML += `⚠️ eBay: ${data.error || 'Check API setup'}<br>`;
                    status.textContent = `eBay Error: ${data.error}`;
                    status.style.color = 'red';
                    exportLog(`eBay warning: ${data.error}`, 'error');
                }
            } catch (e) {
                if (e.name === 'AbortError') {
                    progress.innerHTML += `⚠️ eBay: Request timed out<br>`;
                    status.textContent = 'eBay request timed out';
                    exportLog('eBay request timed out after 2 minutes', 'error');
                } else {
                    progress.innerHTML += `⚠️ eBay: ${e.message}<br>`;
                    status.textContent = `eBay Error: ${e.message}`;
                    exportLog(`eBay error: ${e.message}`, 'error');
                }
                status.style.color = 'red';
            } finally {
                btn.disabled = false;
                btn.innerHTML = '🛒 Publish to eBay (with Ads)';
            }
        }
        
        async function exportEtsyFile() {
            const btn = document.getElementById('btn-export-etsy');
            const status = document.getElementById('export-all-status');
            btn.disabled = true;
            btn.innerHTML = '⏳ Generating...';
            exportLog('Generating Etsy shop uploader...');
            
            try {
                const resp = await fetch('/api/export/etsy', {method: 'POST'});
                const data = await resp.json();
                
                if (data.success) {
                    status.textContent = '✓ Etsy file generated!';
                    status.style.color = 'green';
                    exportLog(`Etsy shop uploader generated: ${data.message}`, 'success');
                } else {
                    status.textContent = `Error: ${data.error}`;
                    status.style.color = 'red';
                    exportLog(`Etsy error: ${data.error}`, 'error');
                }
            } catch (e) {
                status.textContent = 'Error generating Etsy file';
                status.style.color = 'red';
                exportLog(`Error: ${e.message}`, 'error');
            } finally {
                btn.disabled = false;
                btn.innerHTML = '🧡 Generate Etsy File';
            }
        }
        
        // Step 2: Export to all marketplaces
        async function exportAllMarketplaces() {
            const btn = document.getElementById('btn-export-all');
            const status = document.getElementById('export-all-status');
            const progress = document.getElementById('export-progress');
            btn.disabled = true;
            btn.innerHTML = '⏳ Exporting...';
            status.textContent = '';
            progress.style.display = 'block';
            progress.innerHTML = '';
            exportLog('Starting export to all marketplaces...');
            
            try {
                // Get theme and use cases from Generate tab
                const theme = document.getElementById('theme')?.value || '';
                const useCases = document.getElementById('use-cases')?.value || '';
                
                // Step 2a: Amazon Flatfile
                exportLog('Generating Amazon flatfile...');
                progress.innerHTML += '📦 Amazon Flatfile...<br>';
                const amazonResp = await fetch('/api/generate/amazon-flatfile', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({ theme, use_cases: useCases })
                });
                const amazonData = await amazonResp.json();
                if (amazonData.success) {
                    progress.innerHTML += `✅ Amazon: ${amazonData.message}<br>`;
                    exportLog('Amazon flatfile generated', 'success');
                } else {
                    throw new Error(`Amazon: ${amazonData.error}`);
                }
                
                // Step 2b: eBay API (with ads) - with timeout handling
                exportLog('Publishing to eBay (this may take a minute)...');
                exportLog('eBay: Connecting to eBay API...');
                progress.innerHTML += '🛒 eBay API (may take up to 2 minutes)...<br>';
                try {
                    const ebayController = new AbortController();
                    const ebayTimeout = setTimeout(() => ebayController.abort(), 120000); // 2 min timeout
                    
                    exportLog('eBay: Sending listing request with ads enabled...');
                    const ebayResp = await fetch('/api/ebay/publish', {
                        method: 'POST',
                        headers: {'Content-Type': 'application/json'},
                        body: JSON.stringify({ with_ads: true }),
                        signal: ebayController.signal
                    });
                    clearTimeout(ebayTimeout);
                    
                    exportLog(`eBay: Response received (status ${ebayResp.status})`);
                    const ebayData = await ebayResp.json();
                    if (ebayData.success) {
                        progress.innerHTML += `✅ eBay: ${ebayData.count || 0} listings published<br>`;
                        exportLog(`eBay: Listing ID: ${ebayData.listing_id}`, 'success');
                        if (ebayData.url) exportLog(`eBay: URL: ${ebayData.url}`, 'success');
                        if (ebayData.promoted) exportLog('eBay: Ads enabled at 5% rate', 'success');
                        exportLog('eBay listings published', 'success');
                    } else {
                        progress.innerHTML += `⚠️ eBay: ${ebayData.error || 'Check API setup'}<br>`;
                        exportLog(`eBay warning: ${ebayData.error}`, 'error');
                    }
                } catch (ebayErr) {
                    if (ebayErr.name === 'AbortError') {
                        progress.innerHTML += `⚠️ eBay: Request timed out (API may still be processing)<br>`;
                        exportLog('eBay request timed out after 2 minutes', 'error');
                    } else {
                        progress.innerHTML += `⚠️ eBay: ${ebayErr.message}<br>`;
                        exportLog(`eBay error: ${ebayErr.message}`, 'error');
                    }
                }
                
                // Step 2c: Etsy Shop Uploader
                exportLog('Generating Etsy shop uploader...');
                progress.innerHTML += '🧡 Etsy Shop Uploader...<br>';
                const etsyResp = await fetch('/api/export/etsy', {method: 'POST'});
                const etsyData = await etsyResp.json();
                if (etsyData.success) {
                    progress.innerHTML += `✅ Etsy: ${etsyData.message}<br>`;
                    exportLog('Etsy shop uploader generated', 'success');
                } else {
                    progress.innerHTML += `⚠️ Etsy: ${etsyData.error || 'Error'}<br>`;
                    exportLog(`Etsy warning: ${etsyData.error}`, 'error');
                }
                
                progress.innerHTML += '<br><strong>Export complete!</strong>';
                status.textContent = '✓ Complete!';
                status.style.color = 'green';
                exportLog('All marketplace exports complete', 'success');
                
            } catch (e) {
                status.textContent = 'Error - see console';
                status.style.color = 'red';
                progress.innerHTML += `❌ Error: ${e.message}<br>`;
                exportLog(`Error: ${e.message}`, 'error');
            } finally {
                btn.disabled = false;
                btn.innerHTML = '🚀 Export to All Marketplaces';
            }
        }
        
        // Step 3: Download M Number folders as ZIP
        async function downloadMNumberFolders() {
            const btn = document.getElementById('btn-download-folders');
            const status = document.getElementById('folders-status');
            btn.disabled = true;
            btn.innerHTML = '⏳ Generating ZIP...';
            status.textContent = '';
            exportLog('Generating M Number folders ZIP...');
            
            try {
                const resp = await fetch('/api/export/m-folders', {method: 'POST'});
                const data = await resp.json();
                
                if (data.success) {
                    // Trigger download
                    window.location.href = `/api/export/m-folders/download?file=${encodeURIComponent(data.file_path)}`;
                    status.textContent = '✓ Download started!';
                    status.style.color = 'green';
                    exportLog(`ZIP created: ${data.file_name}`, 'success');
                } else {
                    status.textContent = `Error: ${data.error}`;
                    status.style.color = 'red';
                    exportLog(`Error: ${data.error}`, 'error');
                }
            } catch (e) {
                status.textContent = 'Error generating ZIP';
                status.style.color = 'red';
                exportLog(`Error: ${e.message}`, 'error');
            } finally {
                btn.disabled = false;
                btn.innerHTML = '📥 Download M Number Folders (ZIP)';
            }
        }
        
        // Load Amazon flatfile preview table
        async function loadFlatfilePreview() {
            exportLog('Loading flatfile preview...');
            const thead = document.getElementById('flatfile-thead');
            const tbody = document.getElementById('flatfile-tbody');
            
            try {
                const resp = await fetch('/api/export/flatfile-preview');
                const data = await resp.json();
                
                if (data.success) {
                    // Build header row with dark background
                    const headers = data.headers;
                    thead.innerHTML = '<tr>' + headers.map(h => `<th style="padding: 8px 12px; border: 1px solid #555; background: #343a40; color: white; font-weight: bold; min-width: 120px;">${h}</th>`).join('') + '</tr>';
                    
                    // Build data rows - show full values
                    let rowsHtml = '';
                    data.rows.forEach((row, idx) => {
                        const bgColor = idx % 2 === 0 ? '#fff' : '#f0f0f0';
                        const isParent = row.parent_child === 'Parent';
                        const rowStyle = isParent ? 'background: #fff3cd; font-weight: bold;' : `background: ${bgColor};`;
                        rowsHtml += '<tr style="' + rowStyle + '">';
                        headers.forEach(h => {
                            let val = row[h] || '';
                            // Don't truncate - show full values with scroll
                            rowsHtml += `<td style="padding: 6px 8px; border: 1px solid #ddd; max-width: 300px; overflow: hidden; text-overflow: ellipsis;" title="${val}">${val}</td>`;
                        });
                        rowsHtml += '</tr>';
                    });
                    tbody.innerHTML = rowsHtml;
                    
                    exportLog(`Loaded ${data.rows.length} rows`, 'success');
                } else {
                    tbody.innerHTML = `<tr><td colspan="10" style="color: red; padding: 20px;">Error: ${data.error}</td></tr>`;
                    exportLog(`Error: ${data.error}`, 'error');
                }
            } catch (e) {
                tbody.innerHTML = `<tr><td colspan="10" style="color: red; padding: 20px;">Error loading preview</td></tr>`;
                exportLog(`Error: ${e.message}`, 'error');
            }
        }
        
        // Download Amazon flatfile
        async function downloadAmazonFlatfile() {
            exportLog('Generating Amazon flatfile for download...');
            const theme = document.getElementById('theme')?.value || '';
            const useCases = document.getElementById('use-cases')?.value || '';
            
            try {
                const resp = await fetch('/api/export/amazon-flatfile-download', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({ theme, use_cases: useCases })
                });
                
                if (resp.ok) {
                    const blob = await resp.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `amazon_flatfile_${new Date().toISOString().slice(0,10)}.xlsx`;
                    document.body.appendChild(a);
                    a.click();
                    a.remove();
                    window.URL.revokeObjectURL(url);
                    exportLog('Amazon flatfile downloaded', 'success');
                } else {
                    const data = await resp.json();
                    exportLog(`Error: ${data.error}`, 'error');
                }
            } catch (e) {
                exportLog(`Error: ${e.message}`, 'error');
            }
        }
        
        // Download Etsy file
        async function downloadEtsyFile() {
            exportLog('Generating Etsy file for download...');
            
            try {
                const resp = await fetch('/api/export/etsy-download', {method: 'POST'});
                
                if (resp.ok) {
                    const blob = await resp.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `etsy_shop_uploader_${new Date().toISOString().slice(0,10)}.xlsx`;
                    document.body.appendChild(a);
                    a.click();
                    a.remove();
                    window.URL.revokeObjectURL(url);
                    exportLog('Etsy file downloaded', 'success');
                } else {
                    const data = await resp.json();
                    exportLog(`Error: ${data.error}`, 'error');
                }
            } catch (e) {
                exportLog(`Error: ${e.message}`, 'error');
            }
        }
        
        // Load products on page load
        loadProducts();
    </script>
</body>
</html>
'''


@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route('/api/products', methods=['GET'])
def get_products():
    return jsonify(Product.all())


@app.route('/api/products', methods=['POST'])
def create_product():
    data = request.json
    Product.create(data)
    return jsonify({"success": True})


@app.route('/api/products/<m_number>', methods=['GET'])
def get_product(m_number):
    product = Product.get(m_number)
    if product:
        return jsonify(product)
    return jsonify({"error": "Not found"}), 404


@app.route('/api/products/<m_number>', methods=['PATCH'])
def update_product(m_number):
    data = request.json
    Product.update(m_number, data)
    return jsonify({"success": True})


@app.route('/api/products/<m_number>', methods=['DELETE'])
def delete_product(m_number):
    Product.delete(m_number)
    return jsonify({"success": True})


@app.route('/api/preview/<m_number>')
def preview_product(m_number):
    """Generate PNG preview for a product."""
    product = Product.get(m_number)
    if not product:
        return "Not found", 404
    
    try:
        from image_generator import generate_product_image
        png_bytes = generate_product_image(product, "main")
        return Response(png_bytes, mimetype='image/png')
    except Exception as e:
        # Fallback to placeholder SVG on error
        svg = f'''<svg xmlns="http://www.w3.org/2000/svg" width="300" height="200" viewBox="0 0 300 200">
            <rect width="300" height="200" fill="#f8d7da"/>
            <text x="150" y="90" text-anchor="middle" font-family="Arial" font-size="16" fill="#721c24">{m_number}</text>
            <text x="150" y="120" text-anchor="middle" font-family="Arial" font-size="12" fill="#721c24">Preview error</text>
            <text x="150" y="145" text-anchor="middle" font-family="Arial" font-size="10" fill="#999">{str(e)[:40]}</text>
        </svg>'''
        return Response(svg, mimetype='image/svg+xml')


@app.route('/api/analyze/products', methods=['POST'])
def analyze_products():
    """Analyze product images with AI to auto-populate theme and use cases."""
    import os
    import base64
    import logging
    from image_generator import generate_product_image
    
    data = request.json or {}
    sample_m_numbers = data.get('sample_m_numbers', [])
    
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return jsonify({"success": False, "error": "OPENAI_API_KEY not set"}), 400
    
    if not sample_m_numbers:
        return jsonify({"success": False, "error": "No sample images provided"}), 400
    
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        # Build message content with images
        content = []
        
        # Add sample images
        for m_number in sample_m_numbers[:5]:
            product = Product.get(m_number)
            if product:
                try:
                    png_bytes = generate_product_image(product, "main")
                    img_base64 = base64.b64encode(png_bytes).decode()
                    content.append({
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{img_base64}"
                        }
                    })
                except Exception as e:
                    logging.warning(f"Failed to generate sample image for {m_number}: {e}")
        
        # Add analysis prompt
        content.append({
            "type": "text",
            "text": """Analyze these product images. They are self-adhesive aluminum signs.

Please provide:
1. THEME: A brief description of what type of sign this is (e.g., "No entry without permission sign", "Fire exit sign", "Warning sign"). Be specific about what the sign communicates.

2. USE_CASES: Where would this sign typically be used? List 2-4 locations or scenarios, separated by commas (e.g., "offices, warehouses, restricted areas, private property").

Respond in this exact format:
THEME: [your theme description]
USE_CASES: [comma-separated list of use cases]"""
        })
        
        # Call OpenAI GPT-4 Vision API
        response = client.chat.completions.create(
            model="gpt-4o",
            max_tokens=500,
            messages=[
                {"role": "system", "content": "You are a product analyst. Analyze the sign images and identify what they communicate and where they would be used. Be concise and specific."},
                {"role": "user", "content": content}
            ]
        )
        
        result = response.choices[0].message.content
        
        # Parse the response
        theme = ""
        use_cases = ""
        
        for line in result.split('\n'):
            line = line.strip()
            if line.upper().startswith('THEME:'):
                theme = line[6:].strip()
            elif line.upper().startswith('USE_CASES:'):
                use_cases = line[10:].strip()
        
        return jsonify({
            "success": True,
            "theme": theme,
            "use_cases": use_cases,
            "raw_response": result
        })
        
    except Exception as e:
        logging.error(f"Failed to analyze products: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/generate/images', methods=['POST'])
def generate_images():
    """Generate product images for approved products."""
    from image_generator import generate_images_job
    
    products = Product.approved()
    if not products:
        products = Product.all()  # Fall back to all if none approved
    
    if not products:
        return jsonify({"error": "No products to generate"}), 400
    
    job_id = submit_job(
        f"Generate images for {len(products)} products",
        generate_images_job,
        products,
        upload_to_r2=True
    )
    
    return jsonify({"success": True, "job_id": job_id, "count": len(products)})


@app.route('/api/generate/amazon-flatfile', methods=['POST'])
def generate_amazon_flatfile():
    """Generate Amazon flatfile XLSX in proper Amazon format."""
    import logging
    import os
    import re
    import json
    from datetime import datetime
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    data = request.json or {}
    theme = data.get('theme', '')
    use_cases = data.get('use_cases', '')
    ai_content = data.get('ai_content', '')
    
    all_products = Product.all()
    if not all_products:
        return jsonify({"success": False, "error": "No products found"}), 400
    
    # Size dimensions and mappings
    SIZE_DIMENSIONS_CM = {
        "dracula": (9.5, 9.5),
        "saville": (11.0, 9.5),
        "dick": (14.0, 9.0),
        "barzan": (19.0, 14.0),
        "baby_jesus": (29.0, 19.0),
    }
    SIZE_MAP_VALUES = {
        "dracula": "XS",
        "saville": "S",
        "dick": "M",
        "barzan": "L",
        "baby_jesus": "XL",
    }
    SIZE_PRICING = {
        "dracula": 10.99,
        "saville": 11.99,
        "dick": 12.99,
        "barzan": 15.99,
        "baby_jesus": 17.99,
    }
    COLOR_DISPLAY = {"silver": "Silver", "white": "White", "gold": "Gold"}
    
    # R2 public URL base
    R2_PUBLIC_URL = os.environ.get("R2_PUBLIC_URL", "https://pub-f0f96448c91147489e7b6c6b22ed5010.r2.dev")
    
    try:
        # Parse AI content to extract bullet points and description
        # AI content format expected to have titles, bullets, descriptions
        default_bullets = [
            "Premium 1mm brushed aluminium construction with elegant finish provides professional appearance and long-lasting durability",
            "UV-resistant printing technology ensures text remains clear and legible for years, even under harsh sunlight exposure",
            "Self-adhesive backing allows quick peel and stick installation – NO drilling or tools required, ready to use immediately",
            "Fully weatherproof design withstands rain, snow, and temperature extremes, suitable for indoor and outdoor applications",
            "Clear, bold messaging ensures excellent visibility and compliance, perfect for maintaining security in restricted areas"
        ]
        default_description = f"{theme} Sign – Brushed Aluminium, Weatherproof, Self-Adhesive. Professional quality signage for {use_cases}."
        default_search_terms = "sign warning notice metal plaque door wall sticker business office industrial safety weatherproof aluminium"
        
        # Derive parent SKU from theme
        parent_sku = theme.upper().replace(" ", "_").replace("-", "_")
        parent_sku = re.sub(r'[^A-Z0-9_]', '', parent_sku)
        if not parent_sku.endswith("_PARENT"):
            parent_sku = f"{parent_sku}_PARENT"
        
        # Amazon columns
        AMAZON_COLUMNS = [
            ("feed_product_type", "Product Type"),
            ("item_sku", "Seller SKU"),
            ("update_delete", "Update Delete"),
            ("brand_name", "Brand Name"),
            ("external_product_id", "Product ID"),
            ("external_product_id_type", "Product ID Type"),
            ("product_description", "Product Description"),
            ("part_number", "Manufacturer Part Number"),
            ("manufacturer", "Manufacturer"),
            ("item_name", "Item Name (aka Title)"),
            ("language_value", "Language"),
            ("recommended_browse_nodes", "Recommended Browse Nodes"),
            ("main_image_url", "Main Image URL"),
            ("other_image_url1", "Other Image Url1"),
            ("other_image_url2", "Other Image Url2"),
            ("other_image_url3", "Other Image Url3"),
            ("other_image_url4", "Other Image Url4"),
            ("other_image_url5", "Other Image Url5"),
            ("other_image_url6", "Other Image Url6"),
            ("other_image_url7", "Other Image Url7"),
            ("other_image_url8", "Other Image Url8"),
            ("relationship_type", "Relationship Type"),
            ("variation_theme", "Variation Theme"),
            ("parent_sku", "Parent SKU"),
            ("parent_child", "Parentage"),
            ("style_name", "Style Name"),
            ("bullet_point1", "Key Product Features"),
            ("bullet_point2", "Key Product Features"),
            ("bullet_point3", "Key Product Features"),
            ("bullet_point4", "Key Product Features"),
            ("bullet_point5", "Key Product Features"),
            ("generic_keywords", "Search Terms"),
            ("color_name", "Colour"),
            ("size_name", "Size"),
            ("color_map", "Colour Map"),
            ("size_map", "Size Map"),
            ("length_longer_edge", "Item Length Longer Edge"),
            ("length_longer_edge_unit_of_measure", "Item Length Unit"),
            ("width_shorter_edge", "Item Width Shorter Edge"),
            ("width_shorter_edge_unit_of_measure", "Item Width Unit"),
            ("batteries_required", "Batteries Required"),
            ("supplier_declared_dg_hz_regulation5", "Dangerous Goods"),
            ("country_of_origin", "Country/Region Of Origin"),
            ("list_price_with_tax", "List Price"),
            ("merchant_shipping_group_name", "Shipping Group"),
            ("condition_type", "Condition"),
            ("fulfillment_availability#1.fulfillment_channel_code", "Fulfillment Channel"),
            ("fulfillment_availability#1.quantity", "Quantity"),
            ("purchasable_offer[marketplace_id=A1F83G8C2ARO7P]#1.our_price#1.schedule#1.value_with_tax", "Price"),
        ]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template"
        
        # Row 1: Template metadata
        ws.cell(row=1, column=1, value="TemplateType=fptcustom")
        ws.cell(row=1, column=2, value="Version=2025.1207")
        ws.cell(row=1, column=3, value="TemplateSignature=U0lHTkFHRQ==")
        
        # Row 2: Labels
        for col, (attr, label) in enumerate(AMAZON_COLUMNS, 1):
            ws.cell(row=2, column=col, value=label)
        
        # Row 3: Attribute names
        for col, (attr, label) in enumerate(AMAZON_COLUMNS, 1):
            ws.cell(row=3, column=col, value=attr)
        
        # Row 4: Parent row
        parent_title = f"{theme} Sign – Brushed Aluminium, Weatherproof, Self-Adhesive"
        parent_data = {
            "feed_product_type": "signage",
            "item_sku": parent_sku,
            "update_delete": "Update",
            "brand_name": "NorthByNorthEast",
            "product_description": default_description,
            "part_number": parent_sku,
            "item_name": parent_title,
            "recommended_browse_nodes": "330215031",
            "variation_theme": "Size & Colour",
            "parent_child": "Parent",
            "generic_keywords": default_search_terms,
            "batteries_required": "No",
            "supplier_declared_dg_hz_regulation5": "Not Applicable",
            "country_of_origin": "Great Britain",
        }
        for col, (attr, _) in enumerate(AMAZON_COLUMNS, 1):
            ws.cell(row=4, column=col, value=parent_data.get(attr, ""))
        
        # Row 5+: Child products
        row_num = 5
        for product in all_products:
            m_number = product['m_number']
            size = product.get('size', 'saville').lower()
            color = product.get('color', 'silver').lower()
            ean = product.get('ean', '')
            
            dims = SIZE_DIMENSIONS_CM.get(size, (11.0, 9.5))
            size_code = SIZE_MAP_VALUES.get(size, "M")
            price = SIZE_PRICING.get(size, 12.99)
            color_display = COLOR_DISPLAY.get(color, color.title())
            
            # Build title with dimensions
            title = f"{theme} Sign – {dims[0]}x{dims[1]}cm Brushed Aluminium, Weatherproof, Self-Adhesive"
            
            # Style name format: Color_SizeCode
            style_name = f"{color_display}_{size_code}"
            
            # Image URLs
            main_image = f"{R2_PUBLIC_URL}/{m_number} - 001.png"
            
            row_data = {
                "feed_product_type": "signage",
                "item_sku": m_number,
                "update_delete": "Update",
                "brand_name": "NorthByNorthEast",
                "external_product_id": ean,
                "external_product_id_type": "EAN" if ean else "",
                "product_description": default_description,
                "part_number": m_number,
                "manufacturer": "North By North East Print and Sign Limited",
                "item_name": title,
                "language_value": "en_GB",
                "recommended_browse_nodes": "330215031",
                "main_image_url": main_image,
                "other_image_url1": f"{R2_PUBLIC_URL}/{m_number} - 002.png",
                "other_image_url2": f"{R2_PUBLIC_URL}/{m_number} - 003.png",
                "other_image_url3": f"{R2_PUBLIC_URL}/{m_number} - 004.png",
                "other_image_url4": f"{R2_PUBLIC_URL}/{m_number} - 005.png",
                "other_image_url5": f"{R2_PUBLIC_URL}/{m_number} - 006.jpg",
                "relationship_type": "Variation",
                "variation_theme": "Size & Colour",
                "parent_sku": parent_sku,
                "parent_child": "Child",
                "style_name": style_name,
                "bullet_point1": default_bullets[0],
                "bullet_point2": default_bullets[1],
                "bullet_point3": default_bullets[2],
                "bullet_point4": default_bullets[3],
                "bullet_point5": default_bullets[4],
                "generic_keywords": default_search_terms,
                "color_name": color_display,
                "size_name": size_code,
                "color_map": color_display,
                "size_map": size_code,
                "length_longer_edge": str(dims[0]),
                "length_longer_edge_unit_of_measure": "Centimetres",
                "width_shorter_edge": str(dims[1]),
                "width_shorter_edge_unit_of_measure": "Centimetres",
                "batteries_required": "No",
                "supplier_declared_dg_hz_regulation5": "Not Applicable",
                "country_of_origin": "Great Britain",
                "list_price_with_tax": str(price),
                "merchant_shipping_group_name": "RM Tracked 48 Free, 24 -- £2.99, SD -- £7.99",
                "condition_type": "New",
                "fulfillment_availability#1.fulfillment_channel_code": "AMAZON_UK_RAFN",
                "fulfillment_availability#1.quantity": "5",
                "purchasable_offer[marketplace_id=A1F83G8C2ARO7P]#1.our_price#1.schedule#1.value_with_tax": str(price),
            }
            
            for col, (attr, _) in enumerate(AMAZON_COLUMNS, 1):
                ws.cell(row=row_num, column=col, value=row_data.get(attr, ""))
            row_num += 1
        
        # Auto-adjust column widths
        for col in range(1, len(AMAZON_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save to file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = Path(__file__).parent / f"amazon_flatfile_{timestamp}.xlsx"
        wb.save(output_path)
        
        logging.info(f"Amazon flatfile saved to {output_path} with {len(all_products)} products")
        
        return jsonify({
            "success": True,
            "product_count": len(all_products),
            "file_path": str(output_path),
            "message": f"Amazon flatfile saved: {output_path.name}"
        })
        
    except Exception as e:
        logging.error(f"Failed to generate Amazon flatfile: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/jobs')
def list_jobs():
    """List all background jobs."""
    jobs = get_all_jobs()
    return jsonify([job_to_dict(j) for j in jobs])


@app.route('/api/jobs/<job_id>')
def get_job_status(job_id):
    """Get status of a specific job."""
    job = get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify(job_to_dict(job))


@app.route('/api/generate/content', methods=['POST'])
def generate_content():
    """Generate AI content for products using OpenAI GPT-4 with sample images."""
    import os
    import base64
    import logging
    from image_generator import generate_product_image
    
    data = request.json or {}
    theme = data.get('theme', '')
    use_cases = data.get('use_cases', '')
    system_prompt = data.get('system_prompt', '')
    sample_m_numbers = data.get('sample_m_numbers', [])
    
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return jsonify({"success": False, "error": "OPENAI_API_KEY not set"}), 400
    
    all_products = Product.all()
    if not all_products:
        return jsonify({"success": False, "error": "No products found"}), 400
    
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        # Build product summary
        sizes = {}
        colors = set()
        descriptions = set()
        
        for p in all_products:
            size = p.get('size', 'unknown')
            color = p.get('color', 'unknown')
            sizes[size] = sizes.get(size, 0) + 1
            colors.add(color)
            if p.get('description'):
                descriptions.add(p['description'])
        
        size_dims = {
            'saville': '115x95mm (rectangular)',
            'dick': '140x90mm (rectangular)', 
            'barzan': '194x143mm (rectangular)',
            'dracula': '95mm diameter (circular)',
            'baby_jesus': '290x190mm (rectangular)'
        }
        
        summary = f"Total Products: {len(all_products)}\n"
        summary += f"Colors: {', '.join(colors)}\n"
        summary += "Sizes:\n"
        for size, count in sizes.items():
            dims = size_dims.get(size, 'unknown dimensions')
            summary += f"  - {size}: {count} products - {dims}\n"
        summary += f"Product Types: {len(descriptions)} unique designs"
        
        # Build message content with images for GPT-4 Vision
        content = []
        
        # Add sample images
        for m_number in sample_m_numbers[:5]:  # Limit to 5 images
            product = Product.get(m_number)
            if product:
                try:
                    png_bytes = generate_product_image(product, "main")
                    img_base64 = base64.b64encode(png_bytes).decode()
                    content.append({
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{img_base64}"
                        }
                    })
                    logging.info(f"Added sample image for {m_number}")
                except Exception as e:
                    logging.warning(f"Failed to generate sample image for {m_number}: {e}")
        
        # Add text prompt
        user_prompt = f"""=== PRODUCT SUMMARY ===
{summary}

=== USER INPUT ===
Theme: {theme or '(not specified)'}
Use Cases: {use_cases or '(not specified)'}

Please generate Amazon marketplace content for these products. Include:
1. Product titles (under 200 characters)
2. 5 bullet points per product
3. Product descriptions

Remember: These products come in MULTIPLE sizes and shapes as shown in the images and summary above."""

        content.append({"type": "text", "text": user_prompt})
        
        # Call OpenAI GPT-4 Vision API
        response = client.chat.completions.create(
            model="gpt-4o",
            max_tokens=4096,
            messages=[
                {"role": "system", "content": system_prompt or "You are an expert product content writer for Amazon marketplace listings."},
                {"role": "user", "content": content}
            ]
        )
        
        generated_content = response.choices[0].message.content
        
        return jsonify({
            "success": True,
            "content": generated_content
        })
        
    except Exception as e:
        logging.error(f"Failed to generate content: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/generate/full', methods=['POST'])
def generate_full():
    """Run full pipeline."""
    def stream():
        yield "Full pipeline not yet implemented in web version.\n"
    return Response(stream(), mimetype='text/plain')


@app.route('/api/ebay/publish', methods=['POST'])
def publish_to_ebay():
    """Publish approved products to eBay via API with auto-promotion."""
    import logging
    
    data = request.json or {}
    promote = data.get('with_ads', True)
    ad_rate = data.get('ad_rate', '5.0')
    dry_run = data.get('dry_run', False)
    
    products = Product.approved()
    if not products:
        products = Product.all()
    
    if not products:
        return jsonify({"success": False, "error": "No products to publish"}), 400
    
    try:
        from ebay_api import create_ebay_listing, load_policy_ids
        policy_ids = load_policy_ids()
    except FileNotFoundError as e:
        logging.warning(f"eBay policies not configured: {e}")
        return jsonify({"success": False, "error": "eBay policies not configured. Run ebay_setup_policies.py first."}), 400
    except ImportError as e:
        logging.warning(f"eBay API module not available: {e}")
        return jsonify({"success": False, "error": "eBay API module not available"}), 400
    except Exception as e:
        logging.error(f"Error loading eBay policies: {e}")
        return jsonify({"success": False, "error": f"eBay setup error: {str(e)}"}), 400
    
    try:
        logging.info(f"Publishing {len(products)} products to eBay (promote={promote})")
        listing_id = create_ebay_listing(
            products=products,
            policy_ids=policy_ids,
            promote=promote,
            ad_rate=ad_rate,
            dry_run=dry_run,
        )
        
        if listing_id:
            result = {
                "success": True,
                "listing_id": listing_id,
                "count": len(products),
                "promoted": promote and listing_id != "DRY_RUN",
            }
            if listing_id != "DRY_RUN":
                result["url"] = f"https://www.ebay.co.uk/itm/{listing_id}"
            logging.info(f"eBay listing created: {listing_id}")
            return jsonify(result)
        else:
            return jsonify({"success": False, "error": "Failed to create listing - no listing ID returned"}), 500
    except Exception as e:
        logging.error(f"eBay publish error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/flatfile-preview')
def flatfile_preview():
    """Return Amazon flatfile data as JSON for preview table."""
    import os
    
    all_products = Product.all()
    if not all_products:
        return jsonify({"success": False, "error": "No products found"}), 400
    
    # Size dimensions and mappings
    SIZE_DIMENSIONS_CM = {
        "dracula": (9.5, 9.5),
        "saville": (11.0, 9.5),
        "dick": (14.0, 9.0),
        "barzan": (19.0, 14.0),
        "baby_jesus": (29.0, 19.0),
    }
    SIZE_MAP_VALUES = {
        "dracula": "XS",
        "saville": "S",
        "dick": "M",
        "barzan": "L",
        "baby_jesus": "XL",
    }
    SIZE_PRICING = {
        "dracula": 10.99,
        "saville": 11.99,
        "dick": 12.99,
        "barzan": 15.99,
        "baby_jesus": 17.99,
    }
    COLOR_DISPLAY = {"silver": "Silver", "white": "White", "gold": "Gold"}
    R2_PUBLIC_URL = os.environ.get("R2_PUBLIC_URL", "https://pub-f0f96448c91147489e7b6c6b22ed5010.r2.dev")
    
    # Get theme from first product
    theme = all_products[0].get('description', 'Sign') if all_products else 'Sign'
    parent_sku = theme.upper().replace(" ", "_").replace("-", "_")
    import re
    parent_sku = re.sub(r'[^A-Z0-9_]', '', parent_sku)
    if not parent_sku.endswith("_PARENT"):
        parent_sku = f"{parent_sku}_PARENT"
    
    # Headers for display - ALL columns from flatfile (including lifestyle image)
    headers = ['item_sku', 'parent_child', 'item_name', 'color_name', 'size_name', 'external_product_id', 
               'list_price', 'main_image_url', 'other_image_url1', 'other_image_url2', 'other_image_url3', 
               'other_image_url4', 'other_image_url5', 'bullet_point1', 'bullet_point2', 'generic_keywords']
    
    rows = []
    
    default_bullets = [
        "Premium 1mm brushed aluminium construction",
        "UV-resistant printing technology",
    ]
    
    # Parent row
    parent_title = f"{theme} Sign – Brushed Aluminium, Weatherproof, Self-Adhesive"
    rows.append({
        'item_sku': parent_sku,
        'parent_child': 'Parent',
        'item_name': parent_title,
        'color_name': '',
        'size_name': '',
        'external_product_id': '',
        'list_price': '',
        'main_image_url': '',
        'other_image_url1': '',
        'other_image_url2': '',
        'other_image_url3': '',
        'other_image_url4': '',
        'bullet_point1': '',
        'bullet_point2': '',
        'generic_keywords': ''
    })
    
    # Child rows
    for product in all_products:
        m_number = product['m_number']
        size = product.get('size', 'saville').lower()
        color = product.get('color', 'silver').lower()
        ean = product.get('ean', '')
        
        dims = SIZE_DIMENSIONS_CM.get(size, (11.0, 9.5))
        size_code = SIZE_MAP_VALUES.get(size, "M")
        price = SIZE_PRICING.get(size, 12.99)
        color_display = COLOR_DISPLAY.get(color, color.title())
        
        title = f"{theme} Sign – {dims[0]}x{dims[1]}cm Brushed Aluminium"
        main_image = f"{R2_PUBLIC_URL}/{m_number} - 001.png"
        
        rows.append({
            'item_sku': m_number,
            'parent_child': 'Child',
            'item_name': title,
            'color_name': color_display,
            'size_name': size_code,
            'external_product_id': str(ean) if ean else '',
            'list_price': f"£{price:.2f}",
            'main_image_url': f"{R2_PUBLIC_URL}/{m_number} - 001.png",
            'other_image_url1': f"{R2_PUBLIC_URL}/{m_number} - 002.png",
            'other_image_url2': f"{R2_PUBLIC_URL}/{m_number} - 003.png",
            'other_image_url3': f"{R2_PUBLIC_URL}/{m_number} - 004.png",
            'other_image_url4': f"{R2_PUBLIC_URL}/{m_number} - 005.png",
            'other_image_url5': f"{R2_PUBLIC_URL}/{m_number} - 006.jpg",
            'bullet_point1': default_bullets[0],
            'bullet_point2': default_bullets[1],
            'generic_keywords': 'sign warning notice metal plaque weatherproof'
        })
    
    return jsonify({
        "success": True,
        "headers": headers,
        "rows": rows
    })


@app.route('/api/export/amazon-flatfile-download', methods=['POST'])
def download_amazon_flatfile():
    """Generate and download Amazon flatfile XLSX."""
    import os
    import re
    from io import BytesIO
    from datetime import datetime
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    data = request.json or {}
    theme = data.get('theme', '')
    use_cases = data.get('use_cases', '')
    
    all_products = Product.all()
    if not all_products:
        return jsonify({"success": False, "error": "No products found"}), 400
    
    # Size dimensions and mappings
    SIZE_DIMENSIONS_CM = {
        "dracula": (9.5, 9.5), "saville": (11.0, 9.5), "dick": (14.0, 9.0),
        "barzan": (19.0, 14.0), "baby_jesus": (29.0, 19.0),
    }
    SIZE_MAP_VALUES = {"dracula": "XS", "saville": "S", "dick": "M", "barzan": "L", "baby_jesus": "XL"}
    SIZE_PRICING = {"dracula": 10.99, "saville": 11.99, "dick": 12.99, "barzan": 15.99, "baby_jesus": 17.99}
    COLOR_DISPLAY = {"silver": "Silver", "white": "White", "gold": "Gold"}
    R2_PUBLIC_URL = os.environ.get("R2_PUBLIC_URL", "https://pub-f0f96448c91147489e7b6c6b22ed5010.r2.dev")
    
    if not theme:
        theme = all_products[0].get('description', 'Sign')
    
    parent_sku = re.sub(r'[^A-Z0-9_]', '', theme.upper().replace(" ", "_").replace("-", "_"))
    if not parent_sku.endswith("_PARENT"):
        parent_sku = f"{parent_sku}_PARENT"
    
    default_bullets = [
        "Premium 1mm brushed aluminium construction with elegant finish",
        "UV-resistant printing technology ensures text remains clear and legible",
        "Self-adhesive backing allows quick peel and stick installation",
        "Fully weatherproof design withstands rain, snow, and temperature extremes",
        "Clear, bold messaging ensures excellent visibility and compliance"
    ]
    default_description = f"{theme} Sign – Brushed Aluminium, Weatherproof, Self-Adhesive."
    default_search_terms = "sign warning notice metal plaque weatherproof aluminium"
    
    AMAZON_COLUMNS = [
        ("feed_product_type", "Product Type"), ("item_sku", "Seller SKU"), ("update_delete", "Update Delete"),
        ("brand_name", "Brand Name"), ("external_product_id", "Product ID"), ("external_product_id_type", "Product ID Type"),
        ("product_description", "Product Description"), ("part_number", "Part Number"), ("manufacturer", "Manufacturer"),
        ("item_name", "Item Name"), ("recommended_browse_nodes", "Browse Nodes"),
        ("main_image_url", "Main Image URL"), ("other_image_url1", "Other Image 1"), ("other_image_url2", "Other Image 2"),
        ("other_image_url3", "Other Image 3"), ("other_image_url4", "Other Image 4"),
        ("relationship_type", "Relationship Type"), ("variation_theme", "Variation Theme"),
        ("parent_sku", "Parent SKU"), ("parent_child", "Parentage"), ("style_name", "Style Name"),
        ("bullet_point1", "Bullet 1"), ("bullet_point2", "Bullet 2"), ("bullet_point3", "Bullet 3"),
        ("bullet_point4", "Bullet 4"), ("bullet_point5", "Bullet 5"), ("generic_keywords", "Search Terms"),
        ("color_name", "Colour"), ("size_name", "Size"), ("color_map", "Colour Map"), ("size_map", "Size Map"),
        ("list_price_with_tax", "List Price"), ("country_of_origin", "Country"),
    ]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Headers
    for col, (attr, label) in enumerate(AMAZON_COLUMNS, 1):
        ws.cell(row=1, column=col, value=label)
        ws.cell(row=2, column=col, value=attr)
    
    # Parent row
    parent_data = {"feed_product_type": "signage", "item_sku": parent_sku, "update_delete": "Update",
        "brand_name": "NorthByNorthEast", "item_name": f"{theme} Sign – Brushed Aluminium",
        "variation_theme": "Size & Colour", "parent_child": "Parent", "country_of_origin": "Great Britain"}
    for col, (attr, _) in enumerate(AMAZON_COLUMNS, 1):
        ws.cell(row=3, column=col, value=parent_data.get(attr, ""))
    
    # Child rows
    row_num = 4
    for product in all_products:
        m_number = product['m_number']
        size = product.get('size', 'saville').lower()
        color = product.get('color', 'silver').lower()
        ean = product.get('ean', '')
        dims = SIZE_DIMENSIONS_CM.get(size, (11.0, 9.5))
        size_code = SIZE_MAP_VALUES.get(size, "M")
        price = SIZE_PRICING.get(size, 12.99)
        color_display = COLOR_DISPLAY.get(color, color.title())
        
        row_data = {
            "feed_product_type": "signage", "item_sku": m_number, "update_delete": "Update",
            "brand_name": "NorthByNorthEast", "external_product_id": ean,
            "external_product_id_type": "EAN" if ean else "", "product_description": default_description,
            "part_number": m_number, "manufacturer": "North By North East Print and Sign Limited",
            "item_name": f"{theme} Sign – {dims[0]}x{dims[1]}cm Brushed Aluminium",
            "recommended_browse_nodes": "330215031",
            "main_image_url": f"{R2_PUBLIC_URL}/{m_number} - 001.png",
            "other_image_url1": f"{R2_PUBLIC_URL}/{m_number} - 002.png",
            "other_image_url2": f"{R2_PUBLIC_URL}/{m_number} - 003.png",
            "other_image_url3": f"{R2_PUBLIC_URL}/{m_number} - 004.png",
            "other_image_url4": f"{R2_PUBLIC_URL}/{m_number} - 005.png",
            "relationship_type": "Variation", "variation_theme": "Size & Colour",
            "parent_sku": parent_sku, "parent_child": "Child", "style_name": f"{color_display}_{size_code}",
            "bullet_point1": default_bullets[0], "bullet_point2": default_bullets[1],
            "bullet_point3": default_bullets[2], "bullet_point4": default_bullets[3],
            "bullet_point5": default_bullets[4], "generic_keywords": default_search_terms,
            "color_name": color_display, "size_name": size_code, "color_map": color_display,
            "size_map": size_code, "list_price_with_tax": str(price), "country_of_origin": "Great Britain",
        }
        for col, (attr, _) in enumerate(AMAZON_COLUMNS, 1):
            ws.cell(row=row_num, column=col, value=row_data.get(attr, ""))
        row_num += 1
    
    for col in range(1, len(AMAZON_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'amazon_flatfile_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')


@app.route('/api/export/etsy-download', methods=['POST'])
def download_etsy_file():
    """Generate and download Etsy shop uploader XLSX."""
    from io import BytesIO
    from datetime import datetime
    
    try:
        from export_etsy import generate_etsy_xlsx
        from config import R2_PUBLIC_URL
        
        products = Product.approved()
        if not products:
            products = Product.all()
        
        if not products:
            return jsonify({"success": False, "error": "No products found"}), 400
        
        xlsx_bytes = generate_etsy_xlsx(products, R2_PUBLIC_URL)
        
        return send_file(BytesIO(xlsx_bytes), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f'etsy_shop_uploader_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/amazon', methods=['POST'])
@app.route('/api/export/flatfile', methods=['POST'])
def export_flatfile():
    """Export Amazon flatfile for approved products."""
    from io import BytesIO
    import openpyxl
    
    products = Product.approved()
    if not products:
        products = Product.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Headers
    headers = ['m_number', 'description', 'size', 'color', 'ean', 'qa_status']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, p in enumerate(products, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=p.get(header, ''))
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'amazon_flatfile_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    )


@app.route('/api/export/ebay', methods=['POST'])
def export_ebay():
    """Export eBay File Exchange CSV."""
    from export_ebay import generate_ebay_csv
    from config import R2_PUBLIC_URL
    
    products = Product.approved()
    if not products:
        products = Product.all()
    
    csv_content = generate_ebay_csv(products, R2_PUBLIC_URL)
    
    return Response(
        csv_content,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=ebay_listings_{datetime.now().strftime("%Y%m%d_%H%M")}.csv'}
    )


@app.route('/api/export/etsy', methods=['POST'])
def export_etsy():
    """Export Etsy Shop Uploader XLSX."""
    import logging
    from datetime import datetime
    
    try:
        from export_etsy import generate_etsy_xlsx
        from config import R2_PUBLIC_URL
        
        products = Product.approved()
        if not products:
            products = Product.all()
        
        if not products:
            return jsonify({"success": False, "error": "No products found"}), 400
        
        xlsx_bytes = generate_etsy_xlsx(products, R2_PUBLIC_URL)
        
        # Save to file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = Path(__file__).parent / f"etsy_shop_uploader_{timestamp}.xlsx"
        with open(output_path, 'wb') as f:
            f.write(xlsx_bytes)
        
        return jsonify({
            "success": True,
            "product_count": len(products),
            "file_path": str(output_path),
            "message": f"Etsy file saved: {output_path.name}"
        })
    except Exception as e:
        logging.error(f"Failed to generate Etsy file: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/lifestyle-background', methods=['POST'])
def generate_lifestyle_background():
    """Generate a lifestyle background image using DALL-E."""
    import os
    import logging
    import traceback
    
    try:
        import requests
        from datetime import datetime
        
        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            logging.error("OPENAI_API_KEY not set in environment")
            return jsonify({"success": False, "error": "OPENAI_API_KEY not set. Please set the environment variable."}), 400
        
        # Get theme from first product or request
        data = request.json or {}
        theme = data.get('theme', '')
        
        if not theme:
            products = Product.all()
            if products:
                theme = products[0].get('description', 'safety sign')
        
        logging.info(f"Generating lifestyle background for theme: {theme}")
        
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        prompt = """A minimalist photograph of a plain concrete or painted wall in a modern building.
The wall is completely blank and empty with nothing on it at all. 
Simple industrial or office interior with soft natural lighting from a window.
Focus on texture and clean lines. No objects, no decorations, no artwork, no fixtures on the wall.
Architectural photography style, shallow depth of field, professional quality."""
        
        logging.info("Calling DALL-E 3 API...")
        response = client.images.generate(
            model="dall-e-3",
            prompt=prompt,
            size="1024x1024",
            quality="standard",
            n=1,
        )
        
        image_url = response.data[0].url
        logging.info(f"DALL-E returned image URL: {image_url[:50]}...")
        
        # Download and save the image locally
        img_response = requests.get(image_url, timeout=60)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        bg_path = Path(__file__).parent / f"lifestyle_background_{timestamp}.png"
        with open(bg_path, 'wb') as f:
            f.write(img_response.content)
        
        logging.info(f"Lifestyle background saved to {bg_path}")
        
        return jsonify({
            "success": True,
            "background_url": f"/api/export/lifestyle-background/preview?file={bg_path.name}",
            "file_path": str(bg_path)
        })
        
    except ImportError as e:
        logging.error(f"Import error: {e}")
        return jsonify({"success": False, "error": f"Missing dependency: {e}"}), 500
    except Exception as e:
        logging.error(f"Failed to generate lifestyle background: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/lifestyle-background/preview')
def preview_lifestyle_background():
    """Serve the lifestyle background image."""
    file_name = request.args.get('file')
    if not file_name:
        return "File not specified", 400
    
    file_path = Path(__file__).parent / file_name
    if not file_path.exists():
        return "File not found", 404
    
    return send_file(file_path, mimetype='image/png')


@app.route('/api/export/lifestyle-images', methods=['POST'])
def generate_lifestyle_images():
    """Generate lifestyle images by overlaying product PNGs on the background and upload to R2."""
    import logging
    import os
    from PIL import Image
    from io import BytesIO
    from image_generator import generate_product_image
    
    data = request.json or {}
    background_url = data.get('background_url', '')
    
    logging.info(f"Lifestyle images request - background_url: {background_url}")
    
    if not background_url:
        return jsonify({"success": False, "error": "No background URL provided"}), 400
    
    # Extract file name from URL - handle both /api/export/file?file=X and direct paths
    if 'file=' in background_url:
        file_name = background_url.split('file=')[-1]
    elif background_url.startswith('/'):
        # Direct path like /api/export/file?file=lifestyle_background_xxx.png
        file_name = background_url.split('/')[-1]
    else:
        file_name = background_url
    
    logging.info(f"Extracted file_name: {file_name}")
    bg_path = Path(__file__).parent / file_name
    logging.info(f"Looking for background at: {bg_path}")
    
    if not bg_path.exists():
        # Try to find the most recent lifestyle background
        bg_files = list(Path(__file__).parent.glob("lifestyle_background_*.png"))
        if bg_files:
            bg_path = max(bg_files, key=lambda p: p.stat().st_mtime)
            logging.info(f"Using most recent background: {bg_path}")
        else:
            logging.error(f"Background file not found: {bg_path}")
            return jsonify({"success": False, "error": f"Background file not found: {file_name}"}), 400
    
    products = Product.all()
    if not products:
        return jsonify({"success": False, "error": "No products found"}), 400
    
    R2_PUBLIC_URL = os.environ.get("R2_PUBLIC_URL", "https://pub-f0f96448c91147489e7b6c6b22ed5010.r2.dev")
    
    try:
        # Load background image
        background = Image.open(bg_path).convert('RGBA')
        bg_width, bg_height = background.size
        
        count = 0
        images_data = []
        
        # Try to get R2 upload function
        try:
            from r2_storage import upload_image as upload_to_r2
            can_upload = True
        except ImportError:
            can_upload = False
            logging.warning("R2 upload not available - saving locally only")
        
        for product in products:
            try:
                m_number = product['m_number']
                logging.info(f"Creating lifestyle image for {m_number}...")
                
                # Generate transparent product image (not main - use transparent version)
                from image_generator import generate_transparent_product_image
                png_bytes = generate_transparent_product_image(product)
                logging.info(f"Generated transparent PNG for {m_number}: {len(png_bytes)} bytes")
                product_img = Image.open(BytesIO(png_bytes)).convert('RGBA')
                
                # Resize product to fit nicely (about 40% of background width)
                target_width = int(bg_width * 0.4)
                ratio = target_width / product_img.width
                target_height = int(product_img.height * ratio)
                product_img = product_img.resize((target_width, target_height), Image.Resampling.LANCZOS)
                
                # Create composite - position product in center-right area
                composite = background.copy()
                x_pos = int(bg_width * 0.55 - target_width // 2)
                y_pos = int(bg_height * 0.45 - target_height // 2)
                composite.paste(product_img, (x_pos, y_pos), product_img)
                
                # Convert to RGB for JPEG (smaller file size for R2)
                composite_rgb = composite.convert('RGB')
                
                # Save to bytes
                img_bytes = BytesIO()
                composite_rgb.save(img_bytes, 'JPEG', quality=90)
                img_bytes.seek(0)
                
                # Save locally
                lifestyle_path = Path(__file__).parent / f"{m_number}_lifestyle.jpg"
                with open(lifestyle_path, 'wb') as f:
                    f.write(img_bytes.getvalue())
                
                # Upload to R2 if available
                r2_url = None
                if can_upload:
                    try:
                        img_bytes.seek(0)
                        r2_key = f"{m_number} - 006.jpg"  # Lifestyle image as image 6
                        upload_to_r2(img_bytes.getvalue(), r2_key, content_type='image/jpeg')
                        r2_url = f"{R2_PUBLIC_URL}/{r2_key}"
                        logging.info(f"Uploaded lifestyle image to R2: {r2_key}")
                    except Exception as upload_err:
                        logging.warning(f"Failed to upload {m_number} to R2: {upload_err}")
                
                images_data.append({
                    'm_number': m_number,
                    'url': f"/api/export/lifestyle-preview/{m_number}",
                    'r2_url': r2_url
                })
                count += 1
                
            except Exception as e:
                import traceback
                logging.error(f"Failed to create lifestyle for {product['m_number']}: {e}")
                logging.error(traceback.format_exc())
        
        return jsonify({
            "success": True,
            "count": count,
            "images": images_data,
            "message": f"Created {count} lifestyle images"
        })
        
    except Exception as e:
        logging.error(f"Failed to generate lifestyle images: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/lifestyle-preview/<m_number>')
def preview_lifestyle_image(m_number):
    """Serve a lifestyle image preview."""
    file_path = Path(__file__).parent / f"{m_number}_lifestyle.jpg"
    if not file_path.exists():
        file_path = Path(__file__).parent / f"{m_number}_lifestyle.png"
    if not file_path.exists():
        return "File not found", 404
    return send_file(file_path)


@app.route('/api/export/m-folders', methods=['POST'])
def export_m_folders_json():
    """Generate M Number folders ZIP and return JSON with file path."""
    import logging
    from datetime import datetime
    from export_images import generate_m_number_folder_zip
    
    products = Product.approved()
    if not products:
        products = Product.all()
    
    if not products:
        return jsonify({"success": False, "error": "No products found"}), 400
    
    try:
        zip_bytes = generate_m_number_folder_zip(products)
        
        # Save to file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = Path(__file__).parent / f"m_number_folders_{timestamp}.zip"
        with open(output_path, 'wb') as f:
            f.write(zip_bytes)
        
        return jsonify({
            "success": True,
            "file_path": str(output_path),
            "file_name": output_path.name,
            "product_count": len(products)
        })
        
    except Exception as e:
        logging.error(f"Failed to generate M folders ZIP: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/m-folders/download')
def download_m_folders():
    """Download the generated M folders ZIP file."""
    file_path = request.args.get('file')
    if not file_path:
        return "File not specified", 400
    
    path = Path(file_path)
    if not path.exists():
        return "File not found", 404
    
    return send_file(path, mimetype='application/zip', as_attachment=True, download_name=path.name)


@app.route('/api/export/images/<m_number>', methods=['GET'])
def export_product_images(m_number):
    """Get product image(s). If type param provided, return single PNG. Otherwise return ZIP of all."""
    from io import BytesIO
    from image_generator import generate_product_image
    
    product = Product.get(m_number)
    if not product:
        return jsonify({"error": "Product not found"}), 404
    
    # If type parameter provided, return single image
    img_type = request.args.get('type')
    if img_type:
        valid_types = ['main', 'dimensions', 'peel_and_stick', 'rear']
        if img_type not in valid_types:
            return jsonify({"error": f"Invalid type. Must be one of: {valid_types}"}), 400
        
        try:
            png_bytes = generate_product_image(product, img_type)
            return send_file(
                BytesIO(png_bytes),
                mimetype='image/png',
                download_name=f'{m_number}_{img_type}.png'
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    
    # No type param - return ZIP of all images
    from export_images import generate_single_product_zip
    zip_bytes = generate_single_product_zip(product)
    
    return send_file(
        BytesIO(zip_bytes),
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'{m_number}_images.zip'
    )


@app.route('/api/export/images', methods=['POST'])
def export_all_images():
    """Download all product images as ZIP (approved products)."""
    from export_images import generate_images_zip
    from io import BytesIO
    
    products = Product.approved()
    if not products:
        products = Product.all()
    
    if not products:
        return jsonify({"error": "No products to export"}), 400
    
    zip_bytes = generate_images_zip(products)
    
    return send_file(
        BytesIO(zip_bytes),
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'product_images_{datetime.now().strftime("%Y%m%d_%H%M")}.zip'
    )


@app.route('/api/export/m-number-folders/<m_number>', methods=['GET'])
def export_m_number_folder(m_number):
    """Download M Number folder with full structure for staff (single product)."""
    from export_images import generate_single_m_number_folder_zip
    from io import BytesIO
    
    product = Product.get(m_number)
    if not product:
        return jsonify({"error": "Product not found"}), 404
    
    zip_bytes = generate_single_m_number_folder_zip(product)
    
    return send_file(
        BytesIO(zip_bytes),
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'{m_number}_folder.zip'
    )


@app.route('/api/export/m-number-folders', methods=['POST'])
def export_all_m_number_folders():
    """Download all M Number folders with full structure for staff."""
    from export_images import generate_m_number_folder_zip
    from io import BytesIO
    
    products = Product.approved()
    if not products:
        products = Product.all()
    
    if not products:
        return jsonify({"error": "No products to export"}), 400
    
    zip_bytes = generate_m_number_folder_zip(products)
    
    return send_file(
        BytesIO(zip_bytes),
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'm_number_folders_{datetime.now().strftime("%Y%m%d_%H%M")}.zip'
    )


@app.route('/api/products/<m_number>/scale', methods=['PATCH'])
def update_product_scale(m_number):
    """Update icon_scale and text_scale for a product (QA tuning)."""
    data = request.json
    
    product = Product.get(m_number)
    if not product:
        return jsonify({"error": "Product not found"}), 404
    
    updates = {}
    if 'icon_scale' in data:
        updates['icon_scale'] = float(data['icon_scale'])
    if 'text_scale' in data:
        updates['text_scale'] = float(data['text_scale'])
    
    if updates:
        Product.update(m_number, updates)
    
    return jsonify({"success": True, "updates": updates})


@app.route('/api/products/<m_number>/position', methods=['PATCH'])
def update_product_position(m_number):
    """Update icon_offset_x and icon_offset_y for a product (QA positioning)."""
    data = request.json
    
    product = Product.get(m_number)
    if not product:
        return jsonify({"error": "Product not found"}), 404
    
    updates = {}
    if 'icon_offset_x' in data:
        updates['icon_offset_x'] = float(data['icon_offset_x'])
    if 'icon_offset_y' in data:
        updates['icon_offset_y'] = float(data['icon_offset_y'])
    
    if updates:
        Product.update(m_number, updates)
    
    return jsonify({"success": True, "updates": updates})


if __name__ == '__main__':
    app.run(debug=True, port=5000)
