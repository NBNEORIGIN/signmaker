"""Etsy Shop Uploader export generator.

Generates Etsy Shop Uploader compatible XLSX from product database.
Uses exact template format from working exports.
"""
import io
import logging
from typing import Optional
from pathlib import Path

import openpyxl

# Size mappings
SIZE_CONFIG = {
    "dracula": {"display": "95mm x 95mm", "dims": (9.5, 9.5, 0.1), "price": 10.99},
    "saville": {"display": "110mm x 95mm", "dims": (11.0, 9.5, 0.1), "price": 11.99},
    "dick": {"display": "140mm x 90mm", "dims": (14.0, 9.0, 0.1), "price": 12.99},
    "barzan": {"display": "190mm x 140mm", "dims": (19.0, 14.0, 0.1), "price": 15.99},
    "baby_jesus": {"display": "290mm x 190mm", "dims": (29.0, 19.0, 0.1), "price": 17.99},
}

# Color mappings
COLOR_DISPLAY = {
    "silver": "Silver",
    "gold": "Gold",
    "white": "White",
}

# Shop Uploader configuration - these IDs are from your Etsy account
SHOP_UPLOADER_CONFIG = {
    "category": "Signs (2844)",
    "shipping_profile_id": "Postage 2025 (208230423243)",
    "readiness_state_id": 1402336022581,
    "return_policy_id": "return=true|exchange=true|deadline=30 (1074420280634)",
}

# Full Shop Uploader columns (exact match to template)
SHOP_UPLOADER_COLUMNS = [
    "listing_id", "parent_sku", "sku", "title", "description", "price", "quantity",
    "category", "_primary_color", "_secondary_color", "_occasion", "_holiday",
    "_deprecated_diameter", "_deprecated_dimensions", "_deprecated_fabric", "_deprecated_finish",
    "_deprecated_flavor", "_deprecated_height", "_deprecated_length", "_deprecated_material",
    "_deprecated_pattern", "_deprecated_scent", "_deprecated_size", "_deprecated_style",
    "_deprecated_weight", "_deprecated_width", "_deprecated_device",
    "option1_name", "option1_value", "option2_name", "option2_value",
    "image_1", "image_2", "image_3", "image_4", "image_5", "image_6", "image_7", "image_8", "image_9", "image_10",
    "shipping_profile_id", "readiness_state_id", "return_policy_id",
    "length", "width", "height", "dimensions_unit", "weight", "weight_unit",
    "type", "who_made", "is_made_to_order", "year_made", "is_vintage", "is_supply",
    "is_taxable", "auto_renew", "is_customizable", "is_personalizable",
    "personalization_is_required", "personalization_instructions", "personalization_char_count_max",
    "style_1", "style_2",
    "tag_1", "tag_2", "tag_3", "tag_4", "tag_5", "tag_6", "tag_7", "tag_8", "tag_9", "tag_10", "tag_11", "tag_12", "tag_13",
    "action", "listing_state", "overwrite_images"
]


def generate_etsy_xlsx(products: list[dict], r2_public_url: str = "") -> bytes:
    """
    Generate Etsy Shop Uploader compatible XLSX.
    
    Args:
        products: List of product dicts from database
        r2_public_url: Base URL for R2 images
    
    Returns:
        XLSX file as bytes
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Write headers
    for col, header in enumerate(SHOP_UPLOADER_COLUMNS, 1):
        ws.cell(row=1, column=col, value=header)
    
    row_num = 2
    
    # Group products by parent (description)
    parent_groups = {}
    for product in products:
        desc = product.get("description", "Unknown")
        if desc not in parent_groups:
            parent_groups[desc] = []
        parent_groups[desc].append(product)
    
    for desc, group in parent_groups.items():
        # Use description as parent SKU base
        parent_sku = desc.upper().replace(" ", "_") + "_PARENT"
        
        for product in group:
            m_number = product.get("m_number", "")
            size = product.get("size", "dracula").lower()
            color = product.get("color", "silver").lower()
            
            size_info = SIZE_CONFIG.get(size, SIZE_CONFIG["dracula"])
            color_display = COLOR_DISPLAY.get(color, "Silver")
            dims = size_info["dims"]
            
            # Build title (max 140 chars for Etsy)
            title = f"{desc} Sign – {size_info['display']} Brushed Aluminium, Weatherproof, Self-Adhesive"
            if len(title) > 140:
                title = title[:137] + "..."
            
            # Build description
            description = f"""Clearly mark your property with this professional brushed aluminium sign. Perfect for maintaining control over private areas and restricted access spaces.

Crafted from premium 1mm brushed aluminium with a sophisticated {color_display.lower()} finish, this compact {size_info['display']} sign delivers maximum impact whilst maintaining an elegant appearance. The UV-printed design ensures crystal-clear visibility, making your message unmistakable.

Installation couldn't be easier thanks to the high-quality self-adhesive backing – simply peel and stick with NO drilling required. The weatherproof construction and UV-resistant printing guarantee long-lasting performance in all outdoor conditions, whilst rounded corners provide a professional finish and prevent injury.

Ideal for private driveways, residential property entrances, business areas, and commercial zones. This durable sign offers an effective, maintenance-free solution."""
            
            # Image URLs (URL-encoded)
            images = []
            if r2_public_url:
                for img_num in range(1, 5):
                    img_url = f"{r2_public_url}/{m_number}%20-%20{img_num:03d}.jpg"
                    images.append(img_url)
            
            # Generate tags (max 13)
            base_desc = desc.lower()
            tags = [
                base_desc.replace(" ", " "),
                "sign",
                "aluminium sign",
                "safety sign",
                f"{color_display.lower()} sign",
                "office sign",
                "weatherproof",
                "self adhesive",
                "uk sign",
                "property sign",
                "warning sign",
                "metal sign",
                "professional sign"
            ]
            
            row_data = {
                "listing_id": "",
                "parent_sku": parent_sku,
                "sku": m_number,
                "title": title,
                "description": description,
                "price": size_info["price"],
                "quantity": 999,
                "category": SHOP_UPLOADER_CONFIG["category"],
                "_primary_color": color_display,
                "_secondary_color": "",
                "_occasion": "",
                "_holiday": "",
                "option1_name": "Size",
                "option1_value": size_info["display"],
                "option2_name": "_primary_color",
                "option2_value": color_display,
                "image_1": images[0] if len(images) > 0 else "",
                "image_2": images[1] if len(images) > 1 else "",
                "image_3": images[2] if len(images) > 2 else "",
                "image_4": images[3] if len(images) > 3 else "",
                "image_5": "",
                "image_6": "",
                "image_7": "",
                "image_8": "",
                "image_9": "",
                "image_10": "",
                "shipping_profile_id": SHOP_UPLOADER_CONFIG["shipping_profile_id"],
                "readiness_state_id": SHOP_UPLOADER_CONFIG["readiness_state_id"],
                "return_policy_id": SHOP_UPLOADER_CONFIG["return_policy_id"],
                "length": dims[0],
                "width": dims[1],
                "height": dims[2],
                "dimensions_unit": "cm",
                "weight": 50,
                "weight_unit": "g",
                "type": "physical",
                "who_made": "i_did",
                "is_made_to_order": "TRUE",
                "year_made": "",
                "is_vintage": "FALSE",
                "is_supply": "FALSE",
                "is_taxable": "TRUE",
                "auto_renew": "TRUE",
                "is_customizable": "FALSE",
                "is_personalizable": "FALSE",
                "personalization_is_required": "FALSE",
                "personalization_instructions": "",
                "personalization_char_count_max": "",
                "style_1": "Modern",
                "style_2": "",
                "tag_1": tags[0] if len(tags) > 0 else "",
                "tag_2": tags[1] if len(tags) > 1 else "",
                "tag_3": tags[2] if len(tags) > 2 else "",
                "tag_4": tags[3] if len(tags) > 3 else "",
                "tag_5": tags[4] if len(tags) > 4 else "",
                "tag_6": tags[5] if len(tags) > 5 else "",
                "tag_7": tags[6] if len(tags) > 6 else "",
                "tag_8": tags[7] if len(tags) > 7 else "",
                "tag_9": tags[8] if len(tags) > 8 else "",
                "tag_10": tags[9] if len(tags) > 9 else "",
                "tag_11": tags[10] if len(tags) > 10 else "",
                "tag_12": tags[11] if len(tags) > 11 else "",
                "tag_13": tags[12] if len(tags) > 12 else "",
                "action": "create",
                "listing_state": "draft",
                "overwrite_images": "TRUE",
            }
            
            for col, header in enumerate(SHOP_UPLOADER_COLUMNS, 1):
                ws.cell(row=row_num, column=col, value=row_data.get(header, ""))
            
            row_num += 1
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


if __name__ == "__main__":
    # Test
    test_products = [
        {"m_number": "M1001", "description": "No Entry", "size": "saville", "color": "silver"},
        {"m_number": "M1002", "description": "No Entry", "size": "dick", "color": "gold"},
        {"m_number": "M1003", "description": "Staff Only", "size": "saville", "color": "white"},
    ]
    xlsx_bytes = generate_etsy_xlsx(test_products)
    with open("test_etsy.xlsx", "wb") as f:
        f.write(xlsx_bytes)
    print(f"Generated test_etsy.xlsx ({len(xlsx_bytes)} bytes)")
