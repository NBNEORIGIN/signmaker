"""Import Amazon flatfile data into SignMaker database for testing."""
import sys
from pathlib import Path
import openpyxl

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent))

from models import init_db, Product


def parse_size_from_title(title: str) -> str:
    """Extract size from title dimensions."""
    if "9.5x9.5" in title or "9.5 x 9.5" in title:
        return "dracula"
    elif "11x9.5" in title or "11 x 9.5" in title or "11.0 x 9.5" in title:
        return "saville"
    elif "14x9" in title or "14 x 9" in title:
        return "dick"
    elif "19x14" in title or "19 x 14" in title:
        return "barzan"
    elif "29x19" in title or "29 x 19" in title:
        return "baby_jesus"
    return "dracula"


def parse_color_from_description(desc: str) -> str:
    """Extract color from description."""
    desc_lower = desc.lower()
    if "gold" in desc_lower:
        return "gold"
    elif "white" in desc_lower:
        return "white"
    return "silver"


def import_flatfile(flatfile_path: Path):
    """Import products from Amazon flatfile."""
    print(f"Loading flatfile: {flatfile_path}")
    
    wb = openpyxl.load_workbook(flatfile_path, read_only=True, data_only=True)
    ws = wb.active
    
    # Find header row (row 3 typically has attribute names)
    headers = []
    for row_idx, row in enumerate(ws.iter_rows(max_row=5), 1):
        values = [c.value for c in row]
        if 'item_sku' in values or 'feed_product_type' in values:
            headers = values
            header_row = row_idx
            print(f"Found headers at row {header_row}: {headers[:10]}...")
            break
    
    if not headers:
        # Try to use first row
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        header_row = 1
    
    # Map column indices
    col_map = {h: i for i, h in enumerate(headers) if h}
    
    print(f"Columns found: {list(col_map.keys())[:15]}...")
    
    # Initialize database
    init_db()
    
    products_imported = 0
    
    # Read data rows
    for row in ws.iter_rows(min_row=header_row + 1):
        values = [c.value for c in row]
        
        # Get M number from item_sku
        sku_idx = col_map.get('item_sku', col_map.get('seller_sku', 1))
        m_number = values[sku_idx] if sku_idx < len(values) else None
        
        if not m_number or not str(m_number).startswith('M'):
            continue
        
        m_number = str(m_number).strip()
        
        # Get EAN
        ean_idx = col_map.get('external_product_id', col_map.get('product_id', 4))
        ean = str(values[ean_idx]) if ean_idx < len(values) and values[ean_idx] else ""
        
        # Get title
        title_idx = col_map.get('item_name', col_map.get('product_name', 9))
        title = values[title_idx] if title_idx < len(values) else ""
        
        # Get description
        desc_idx = col_map.get('product_description', col_map.get('description', 6))
        description = values[desc_idx] if desc_idx < len(values) else ""
        
        # Parse size and color
        size = parse_size_from_title(str(title) if title else "")
        color = parse_color_from_description(str(description) if description else "")
        
        # Get image URLs
        main_img_idx = col_map.get('main_image_url', col_map.get('main-image-url', 11))
        main_image = values[main_img_idx] if main_img_idx < len(values) else ""
        
        # Create product data
        product_data = {
            "m_number": m_number,
            "ean": ean,
            "description": "No Entry Without Permission",  # Base description
            "size": size,
            "color": color,
            "orientation": "landscape",
            "mounting_type": "self_adhesive",
            "layout_mode": "B",
            "icon_files": "No Entry Without Permission.svg",
            "text_line_1": "",
            "text_line_2": "",
            "text_line_3": "",
            "qa_status": "approved",
            "icon_scale": 1.0,
            "text_scale": 1.0,
        }
        
        # Check if product exists
        existing = Product.get(m_number)
        if existing:
            print(f"  Updating {m_number} ({size}/{color})")
            Product.update(m_number, product_data)
        else:
            print(f"  Creating {m_number} ({size}/{color})")
            Product.create(product_data)
        
        products_imported += 1
    
    print(f"\nImported {products_imported} products")
    return products_imported


if __name__ == "__main__":
    flatfile_path = Path(r"G:\My Drive\003 APPS\019 - AMAZON PUBLISHER REV 2.0\003 FLATFILES\NO ENTRY WITHOUT PERMISSION FLATFILE REV1.xlsm")
    
    if not flatfile_path.exists():
        print(f"Flatfile not found: {flatfile_path}")
        sys.exit(1)
    
    import_flatfile(flatfile_path)
